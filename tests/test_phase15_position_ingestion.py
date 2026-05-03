"""Phase 15 — Investment Summary / Account-Position ingestion tests.

15 tests across schema validation, ingestion mechanics, liquidity mapping,
position_terms_status, and discovery. See MODEL_DOCUMENTATION.md
§Phase 15 design.

DISCIPLINE — Phase 15 reviewer tightening (synthetic fixtures only)
=====================================================================

Tests use SYNTHETIC workbook fixtures only. Each test that requires a
workbook constructs a minimal openpyxl workbook in tmp_path via the
openpyxl write API. NO real workbook rows, live values, position names,
manager names, or fund identifiers committed. The real Investment Summary
workbook is used for local validation only.

Schema (4):
1. PositionRecord: market_value_usd >= 0; negative raises.
2. PositionRecord: liquidity_bucket Literal accepted; valuation_date required.
3. ManagerTermsRecord: T5 completeness (non-unknown confidence requires fields).
4. PositionManifestConfig: duplicate account_ids; invalid liquidity_tier_overrides.

AccountRecord (1):
5. synthetic: prefix allowed; non-URL-safe raises; entity_id colon rejected.

Ingestion mechanics (4):
6. Missing workbook path → FileNotFoundError with resolved path.
7. Flat layout → synthetic AccountRecord produced; positions parsed.
8. Valuation date T2 fallback chain (position row → spec → manifest).
9. Stale valuation count increments for positions > 90 days old.

Liquidity mapping (2):
10. Default mapping: all 9 buckets resolve to correct Phase 12 tier.
11. Override: re_stabilized → locked_strategic honored; others unchanged.

position_terms_status (2):
12. missing_terms when manager_id is None.
13. unknown_confidence when confidence='unknown'; complete_terms when
    confidence='contractual' with all required fields populated.

Discovery (2):
14. Draft manifest privacy_safe mode does not emit raw sheet names.
15. Default manifest config byte-stable across repeated model_dump calls.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import pytest
from aa_model.ingestion.liquidity_mapping import (
    build_effective_mapping,
    resolve_phase12_tier,
)
from aa_model.ingestion.schemas_position import (
    AccountRecord,
    AccountSheetSpec,
    ManagerTermsRecord,
    PositionManifestConfig,
    PositionRecord,
)
from pydantic import ValidationError

# ---- synthetic workbook builder --------------------------------------------


def _build_synthetic_position_workbook(
    path: Path,
    *,
    sheets: dict[str, list[list[object]]],
) -> None:
    """Construct a minimal openpyxl workbook at path.

    Synthetic discipline: row labels, names, and amounts are arbitrary
    placeholders. Person names and live identifiers are NEVER used.
    """
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))


def _minimal_manifest(
    accounts: list[AccountSheetSpec] | None = None,
    *,
    as_of_date: datetime.date | None = None,
) -> PositionManifestConfig:
    return PositionManifestConfig(
        manifest_version="1",
        workbook_version="v1",
        expected_filename="synthetic_v1.xlsx",
        as_of_date=as_of_date or datetime.date(2026, 3, 31),
        accounts=accounts or [],
        manager_terms=[],
        liquidity_tier_overrides=None,
    )


# ---- 1-2. PositionRecord schema -------------------------------------------


def test_position_record_negative_nav_raises():
    """Phase 15 #1: market_value_usd < 0 raises."""
    with pytest.raises(ValidationError, match="market_value_usd must be >= 0"):
        PositionRecord(
            position_id="p1",
            account_id="acct_a",
            market_value_usd=-100.0,
            liquidity_bucket="daily_liquid",
            valuation_date=datetime.date(2026, 3, 31),
            source_row=2,
        )


def test_position_record_valid_and_liquidity_enum():
    """Phase 15 #2: valid PositionRecord accepted; all 9 liquidity_bucket values."""
    buckets = [
        "cash_equivalent",
        "daily_liquid",
        "semi_liquid",
        "illiquid",
        "locked_strategic",
        "re_stabilized",
        "re_development",
        "re_land",
        "opco_strategic",
    ]
    for i, bucket in enumerate(buckets):
        pos = PositionRecord(
            position_id=f"p{i}",
            account_id="acct_a",
            market_value_usd=500_000.0,
            liquidity_bucket=bucket,
            valuation_date=datetime.date(2026, 3, 31),
            source_row=i + 2,
        )
        assert pos.liquidity_bucket == bucket

    # valuation_date is required (no default)
    with pytest.raises((ValidationError, TypeError)):
        PositionRecord(
            position_id="p_bad",
            account_id="acct_a",
            market_value_usd=1.0,
            liquidity_bucket="illiquid",
            source_row=1,
        )


# ---- 3. ManagerTermsRecord T5 -------------------------------------------


def test_manager_terms_t5_completeness():
    """Phase 15 #3: T5 completeness validation on ManagerTermsRecord."""
    # Unknown confidence: all fields None allowed.
    mgr = ManagerTermsRecord(manager_id="mgr_a", confidence="unknown")
    assert mgr.redemption_frequency is None

    # Contractual requires redemption_frequency, fee info, source.
    with pytest.raises(ValidationError, match="redemption_frequency"):
        ManagerTermsRecord(
            manager_id="mgr_b",
            confidence="contractual",
            # missing redemption_frequency, fee_basis, source_document
        )

    # Contractual with required fields: valid.
    mgr_ok = ManagerTermsRecord(
        manager_id="mgr_c",
        confidence="contractual",
        redemption_frequency="quarterly",
        fee_basis="nav",
        source_document="LPA 2022-01-15",
    )
    assert mgr_ok.confidence == "contractual"

    # Rate fields must be in [0.0, 1.0].
    with pytest.raises(ValidationError, match="Rate value must be in"):
        ManagerTermsRecord(
            manager_id="mgr_d",
            confidence="unknown",
            gate_pct=1.5,
        )


# ---- 4. PositionManifestConfig validation ----------------------------------


def test_manifest_config_validation():
    """Phase 15 #4: duplicate account_ids and bad liquidity_tier_overrides."""
    spec_a = AccountSheetSpec(
        account_id="acct_1",
        entity_id="entity_a",
        sheet_name="SheetA",
    )
    spec_b = AccountSheetSpec(
        account_id="acct_1",  # duplicate
        entity_id="entity_b",
        sheet_name="SheetB",
    )
    with pytest.raises(ValidationError, match="Duplicate account_id"):
        PositionManifestConfig(
            manifest_version="1",
            workbook_version="v1",
            as_of_date=datetime.date(2026, 3, 31),
            accounts=[spec_a, spec_b],
        )

    # Invalid Phase15 bucket key.
    with pytest.raises(ValidationError, match="not a valid Phase 15"):
        PositionManifestConfig(
            manifest_version="1",
            workbook_version="v1",
            as_of_date=datetime.date(2026, 3, 31),
            liquidity_tier_overrides={"not_a_bucket": "liquid"},
        )

    # Invalid Phase12 tier value.
    with pytest.raises(ValidationError, match="not a valid Phase 12"):
        PositionManifestConfig(
            manifest_version="1",
            workbook_version="v1",
            as_of_date=datetime.date(2026, 3, 31),
            liquidity_tier_overrides={"re_stabilized": "unknown_tier"},
        )


# ---- 5. AccountRecord synthetic prefix -------------------------------------


def test_account_record_synthetic_prefix():
    """Phase 15 #5: synthetic: prefix allowed; colon elsewhere rejected."""
    # synthetic: prefix accepted.
    acct = AccountRecord(
        account_id="synthetic:sheet_003",
        entity_id="entity_a",
        valuation_date=datetime.date(2026, 3, 31),
    )
    assert acct.account_id == "synthetic:sheet_003"

    # Non-URL-safe account_id (not synthetic) rejected.
    with pytest.raises(ValidationError, match="URL-safe"):
        AccountRecord(
            account_id="bad id!",
            entity_id="entity_a",
            valuation_date=datetime.date(2026, 3, 31),
        )

    # entity_id colon rejected.
    with pytest.raises(ValidationError, match="colons"):
        AccountRecord(
            account_id="acct_ok",
            entity_id="entity:bad",
            valuation_date=datetime.date(2026, 3, 31),
        )


# ---- 6-9. Ingestion mechanics ----------------------------------------------


def test_missing_workbook_raises(tmp_path: Path):
    """Phase 15 #6: missing workbook → FileNotFoundError."""
    from aa_model.ingestion.investment_summary import ingest_investment_summary

    manifest = _minimal_manifest()
    with pytest.raises(FileNotFoundError, match="not found"):
        ingest_investment_summary(
            tmp_path / "nonexistent.xlsx",
            manifest,
            manifest_version="1",
        )


def test_flat_layout_produces_account_and_positions(tmp_path: Path):
    """Phase 15 #7: flat layout produces AccountRecord + positions."""
    from aa_model.ingestion.investment_summary import ingest_investment_summary

    wb_path = tmp_path / "synthetic_v1.xlsx"
    _build_synthetic_position_workbook(
        wb_path,
        sheets={
            "Holdings": [
                ["Name", "Market Value"],
                ["Position A", 100_000.0],
                ["Position B", 250_000.0],
                ["Position C", 50_000.0],
            ]
        },
    )

    spec = AccountSheetSpec(
        account_id="acct_001",
        entity_id="entity_a",
        sheet_name="Holdings",
        layout_type="flat_position",
        header_row_index=0,
        value_column_index=1,
        name_column_index=0,
    )
    manifest = _minimal_manifest([spec])

    result = ingest_investment_summary(wb_path, manifest, manifest_version="1")

    assert len(result.accounts) == 1
    assert result.accounts[0].account_id == "acct_001"
    assert len(result.positions) == 3
    assert all(p.account_id == "acct_001" for p in result.positions)
    assert all(p.market_value_usd >= 0 for p in result.positions)


def test_valuation_date_fallback_chain(tmp_path: Path):
    """Phase 15 #8: T2 fallback: spec date preferred over manifest as_of_date."""
    from aa_model.ingestion.investment_summary import ingest_investment_summary

    wb_path = tmp_path / "synthetic_v1.xlsx"
    _build_synthetic_position_workbook(
        wb_path,
        sheets={
            "Holdings": [
                ["Name", "Value"],
                ["Position A", 1_000.0],
            ]
        },
    )

    spec_date = datetime.date(2025, 12, 31)
    spec = AccountSheetSpec(
        account_id="acct_001",
        entity_id="entity_a",
        sheet_name="Holdings",
        layout_type="flat_position",
        header_row_index=0,
        value_column_index=1,
        valuation_date=spec_date,
    )
    manifest = _minimal_manifest(
        [spec],
        as_of_date=datetime.date(2026, 3, 31),
    )

    result = ingest_investment_summary(wb_path, manifest, manifest_version="1")
    assert len(result.positions) == 1
    # Position should use the spec-level date as fallback.
    assert result.positions[0].valuation_date == spec_date
    assert result.diagnostics.positions_with_fallback_valuation_date == 1


def test_stale_valuation_count(tmp_path: Path):
    """Phase 15 #9: stale_valuation_count increments for positions > 90 days old."""
    from aa_model.ingestion.investment_summary import ingest_investment_summary

    wb_path = tmp_path / "synthetic_v1.xlsx"
    _build_synthetic_position_workbook(
        wb_path,
        sheets={
            "Holdings": [
                ["Name", "Value"],
                ["Position A", 1_000.0],
                ["Position B", 2_000.0],
            ]
        },
    )

    # as_of_date is 2026-03-31; spec.valuation_date is 2025-06-30 → 274 days → stale
    spec = AccountSheetSpec(
        account_id="acct_001",
        entity_id="entity_a",
        sheet_name="Holdings",
        layout_type="flat_position",
        header_row_index=0,
        value_column_index=1,
        valuation_date=datetime.date(2025, 6, 30),
    )
    manifest = _minimal_manifest(
        [spec],
        as_of_date=datetime.date(2026, 3, 31),
    )

    result = ingest_investment_summary(wb_path, manifest, manifest_version="1")
    assert result.diagnostics.stale_valuation_count == 2
    assert result.diagnostics.max_valuation_age_days > 90


# ---- 10-11. Liquidity mapping ----------------------------------------------


def test_liquidity_mapping_defaults():
    """Phase 15 #10: all 9 Phase 15 buckets resolve to correct Phase 12 tiers."""
    expected = {
        "cash_equivalent": "liquid",
        "daily_liquid": "liquid",
        "semi_liquid": "semi_liquid",
        "illiquid": "illiquid",
        "locked_strategic": "locked_strategic",
        "re_stabilized": "illiquid",
        "re_development": "locked_strategic",
        "re_land": "locked_strategic",
        "opco_strategic": "locked_strategic",
    }
    for bucket, tier in expected.items():
        assert resolve_phase12_tier(bucket, None) == tier, bucket

    # Unknown bucket raises.
    with pytest.raises(ValueError, match="Unknown Phase 15"):
        resolve_phase12_tier("magic_bucket", None)


def test_liquidity_mapping_override():
    """Phase 15 #11: re_stabilized → locked_strategic override honored."""
    overrides = {"re_stabilized": "locked_strategic"}
    assert resolve_phase12_tier("re_stabilized", overrides) == "locked_strategic"
    # Non-overridden buckets unchanged.
    assert resolve_phase12_tier("daily_liquid", overrides) == "liquid"
    assert resolve_phase12_tier("illiquid", overrides) == "illiquid"

    # Effective mapping reflects the override.
    mapping = build_effective_mapping(overrides)
    assert mapping["re_stabilized"] == "locked_strategic"
    assert mapping["cash_equivalent"] == "liquid"


# ---- 12-13. position_terms_status ------------------------------------------


def test_terms_status_missing_manager(tmp_path: Path):
    """Phase 15 #12: position with no manager_id → missing_terms."""
    from aa_model.ingestion.investment_summary import ingest_investment_summary

    wb_path = tmp_path / "synthetic_v1.xlsx"
    _build_synthetic_position_workbook(
        wb_path,
        sheets={
            "Holdings": [
                ["Name", "Value"],
                ["Position A", 500_000.0],
            ]
        },
    )
    spec = AccountSheetSpec(
        account_id="acct_001",
        entity_id="entity_a",
        sheet_name="Holdings",
        layout_type="flat_position",
        header_row_index=0,
        value_column_index=1,
    )
    manifest = _minimal_manifest([spec])

    result = ingest_investment_summary(wb_path, manifest, manifest_version="1")
    status = result.diagnostics.position_terms_status
    assert status.get("missing_terms", 0) == 1
    assert status.get("complete_terms", 0) == 0


def test_terms_status_complete_and_unknown(tmp_path: Path):
    """Phase 15 #13: unknown_confidence and complete_terms classified correctly."""
    from aa_model.ingestion.investment_summary import ingest_investment_summary

    wb_path = tmp_path / "synthetic_v1.xlsx"
    _build_synthetic_position_workbook(
        wb_path,
        sheets={
            "Holdings": [
                ["Name", "Value", "Manager"],
                ["Position A", 1_000.0, "mgr_unknown"],
                ["Position B", 2_000.0, "mgr_complete"],
            ]
        },
    )
    spec = AccountSheetSpec(
        account_id="acct_001",
        entity_id="entity_a",
        sheet_name="Holdings",
        layout_type="flat_position",
        header_row_index=0,
        value_column_index=1,
        position_column_mappings={"manager_id": 2},
    )
    # Manager with unknown confidence (placeholder).
    mgr_unknown = ManagerTermsRecord(
        manager_id="mgr_unknown",
        confidence="unknown",
    )
    # Manager with contractual confidence + required fields including notice_days.
    # notice_days required because positions default to liquidity_bucket="illiquid".
    mgr_complete = ManagerTermsRecord(
        manager_id="mgr_complete",
        confidence="contractual",
        redemption_frequency="quarterly",
        notice_days=65,
        fee_basis="nav",
        source_document="LPA 2022-01-15",
    )
    manifest = PositionManifestConfig(
        manifest_version="1",
        workbook_version="v1",
        as_of_date=datetime.date(2026, 3, 31),
        accounts=[spec],
        manager_terms=[mgr_unknown, mgr_complete],
    )

    result = ingest_investment_summary(wb_path, manifest, manifest_version="1")
    status = result.diagnostics.position_terms_status
    assert status.get("unknown_confidence", 0) == 1
    assert status.get("complete_terms", 0) == 1


# ---- 14-15. Discovery + byte stability ------------------------------------


def test_discovery_privacy_safe_no_sheet_names(tmp_path: Path):
    """Phase 15 #14: privacy_safe draft manifest does not emit raw sheet names."""
    from aa_model.ingestion.discovery_position import (
        build_draft_position_manifest,
        discover_investment_summary,
    )
    from openpyxl import Workbook

    # Build a synthetic workbook with a recognizable sheet name.
    wb = Workbook()
    ws = wb.active
    ws.title = "SENSITIVE_FUND_NAME_2026"
    ws.append(["Position Name", "Market Value", "Manager"])
    ws.append(["Position X", 100_000, "Mgr A"])
    wb.save(str(tmp_path / "synthetic.xlsx"))

    discovery = discover_investment_summary(tmp_path / "synthetic.xlsx")
    draft = build_draft_position_manifest(
        discovery,
        mode="privacy_safe",
        workbook_version="v1",
    )

    # privacy_safe mode must not contain the real sheet name.
    assert "SENSITIVE_FUND_NAME_2026" not in draft.manifest_yaml_text


def test_manifest_config_byte_stable():
    """Phase 15 #15: PositionManifestConfig.model_dump is deterministic."""
    spec = AccountSheetSpec(
        account_id="acct_001",
        entity_id="entity_a",
        sheet_name="<TODO_sheet>",
        layout_type="flat_position",
        header_row_index=0,
        value_column_index=1,
        name_column_index=0,
    )
    manifest = PositionManifestConfig(
        manifest_version="1",
        workbook_version="v1",
        as_of_date=datetime.date(2026, 3, 31),
        accounts=[spec],
        manager_terms=[],
        liquidity_tier_overrides=None,
    )
    dump1 = manifest.model_dump(mode="json")
    dump2 = manifest.model_dump(mode="json")
    assert dump1 == dump2
