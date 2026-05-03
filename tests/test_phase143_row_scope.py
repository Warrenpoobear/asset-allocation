"""Phase 14.3 — workbook row_range / data-region support.

8 tests. Synthetic workbook only — no live data, no real workbook.
See MODEL_DOCUMENTATION.md §Phase 14.3 design.

Coverage (8 tests):
1. no row_range → full body extracted (byte-stable baseline)
2. row_range set → only rows in range emitted; rows outside absent
3. two specs, same sheet, non-overlapping ranges → both ingest; duplicate
   row labels in separate ranges do not collide
4. two specs, same sheet, overlapping ranges → model_validate raises
5. two specs, same sheet, one spec missing row_range → model_validate raises
6. row_range start ≤ header_row_index → ingestor raises ValueError at runtime
7. row_range end beyond sheet length → no crash; rows within range emitted
8. display_only + row_range → rows skipped without error (range ignored)
"""

from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest

from aa_model.ingestion.schemas import (
    EntitySheetSpec,
    IngestionResult,
    WorkbookManifestConfig,
)
from aa_model.ingestion.workbook import ingest_workbook


# ---- synthetic workbook helpers --------------------------------------------


def _make_workbook(
    sheet_name: str,
    header_row: int,
    *,
    sections: list[list[tuple[str, dict[str, float]]]],
    gap_rows: int = 1,
) -> bytes:
    """Build an in-memory .xlsx with one sheet.

    ``sections`` is a list of sections; each section is a list of
    (row_label, {quarter: amount}) pairs.  Sections are separated by
    ``gap_rows`` blank rows.  The header row is written at ``header_row``
    (1-indexed) using 'Q1 2026', 'Q2 2026', 'Q3 2026' as quarter headers.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    quarters = ["Q1 2026", "Q2 2026", "Q3 2026"]

    # Write header row
    for col_idx, q in enumerate(quarters, start=2):
        ws.cell(row=header_row, column=col_idx, value=q)

    current_row = header_row + 1
    for sec_idx, section in enumerate(sections):
        if sec_idx > 0:
            current_row += gap_rows  # blank gap between sections
        for label, amounts in section:
            ws.cell(row=current_row, column=1, value=label)
            for col_idx, q in enumerate(quarters, start=2):
                val = amounts.get(q)
                if val is not None:
                    ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _manifest(
    sheet_name: str,
    specs: list[dict],
    *,
    header_row_index: int = 1,
) -> WorkbookManifestConfig:
    entity_sheets = []
    for s in specs:
        entity_sheets.append(
            dict(
                sheet_name=sheet_name,
                entity_type="operating_llc",
                display_name=s.get("entity_id", "entity"),
                parent_entity_id=None,
                cash_flow_role="operating",
                row_classification_rules=[],
                header_row_index=header_row_index,
                period_header_format="q_yyyy",
                layout_type=s.get("layout_type", "horizontal_quarter"),
                **{k: v for k, v in s.items() if k not in ("layout_type",)},
            )
        )
    return WorkbookManifestConfig.model_validate(
        {
            "workbook_version": "v_test",
            "expected_workbook_filename": "test.xlsx",
            "family_aggregate_sheets": [],
            "entity_sheets": entity_sheets,
            "re_partnership_sheets": [],
            "board_snapshot_sheets": [],
            "period_header_format": "q_yyyy",
            "default_header_row_index": header_row_index,
            "subtotal_label_patterns": ["total"],
        }
    )


def _run(wb_bytes: bytes, manifest: WorkbookManifestConfig) -> IngestionResult:
    import tempfile
    from pathlib import Path

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(wb_bytes)
        tmp = Path(f.name)
    try:
        return ingest_workbook(tmp, manifest, manifest_version="1")
    finally:
        tmp.unlink(missing_ok=True)


# ---- 1. no row_range → full body byte-stable --------------------------------


def test_no_row_range_full_body():
    """Phase 14.3 #1: no row_range → all body rows extracted."""
    wb = _make_workbook(
        "Sheet1",
        header_row=1,
        sections=[
            [("Alpha", {"Q1 2026": -100.0}), ("Beta", {"Q2 2026": -200.0})],
        ],
    )
    manifest = _manifest("Sheet1", [{"entity_id": "ent_a"}], header_row_index=1)
    result = _run(wb, manifest)
    labels = {r.row_label for r in result.cash_flow_lines}
    assert "Alpha" in labels
    assert "Beta" in labels
    assert len(result.cash_flow_lines) == 2


# ---- 2. row_range → only rows in range emitted ------------------------------


def test_row_range_restricts_to_range():
    """Phase 14.3 #2: row_range=[2,2] emits only row 2; row 3 absent."""
    # header=row1; Alpha=row2; Beta=row3
    wb = _make_workbook(
        "Sheet1",
        header_row=1,
        sections=[
            [("Alpha", {"Q1 2026": -100.0}), ("Beta", {"Q2 2026": -200.0})],
        ],
    )
    manifest = _manifest(
        "Sheet1",
        [{"entity_id": "ent_a", "row_range": [2, 2]}],
        header_row_index=1,
    )
    result = _run(wb, manifest)
    labels = {r.row_label for r in result.cash_flow_lines}
    assert "Alpha" in labels
    assert "Beta" not in labels


# ---- 3. two specs, non-overlapping ranges → no collision --------------------


def test_two_scoped_specs_no_collision():
    """Phase 14.3 #3: duplicate row labels in separate sections ingest cleanly."""
    # header=row1; Section1: Alpha row2, Beta row3
    #              gap row4
    #              Section2: Alpha row5, Beta row6  (same labels!)
    wb = _make_workbook(
        "Sheet1",
        header_row=1,
        sections=[
            [("Alpha", {"Q1 2026": -10.0}), ("Beta", {"Q1 2026": -20.0})],
            [("Alpha", {"Q1 2026": -30.0}), ("Beta", {"Q1 2026": -40.0})],
        ],
        gap_rows=1,
    )
    # rows: 1=header, 2=Alpha(s1), 3=Beta(s1), 4=blank, 5=Alpha(s2), 6=Beta(s2)
    manifest = _manifest(
        "Sheet1",
        [
            {"entity_id": "ent_sec1", "row_range": [2, 3]},
            {"entity_id": "ent_sec2", "row_range": [5, 6]},
        ],
        header_row_index=1,
    )
    result = _run(wb, manifest)
    # Two lines per section × 1 quarter each = 4 total
    assert len(result.cash_flow_lines) == 4
    entity_ids = {r.entity_id for r in result.cash_flow_lines}
    assert entity_ids == {"ent_sec1", "ent_sec2"}
    # Both "Alpha" + "Beta" in both entities
    for eid in ("ent_sec1", "ent_sec2"):
        eid_labels = {r.row_label for r in result.cash_flow_lines if r.entity_id == eid}
        assert eid_labels == {"Alpha", "Beta"}


# ---- 4. overlapping ranges → model_validate raises --------------------------


def test_overlapping_row_ranges_raises():
    """Phase 14.3 #4: overlapping row_range values raise at manifest validation."""
    with pytest.raises(ValueError, match="overlap"):
        _manifest(
            "Sheet1",
            [
                {"entity_id": "ent_a", "row_range": [2, 5]},
                {"entity_id": "ent_b", "row_range": [5, 8]},  # 5 overlaps
            ],
            header_row_index=1,
        )


# ---- 5. one spec missing row_range → model_validate raises ------------------


def test_missing_row_range_on_shared_sheet_raises():
    """Phase 14.3 #5: one spec without row_range on a shared sheet raises."""
    with pytest.raises(ValueError, match="row_range"):
        _manifest(
            "Sheet1",
            [
                {"entity_id": "ent_a", "row_range": [2, 4]},
                {"entity_id": "ent_b"},  # no row_range
            ],
            header_row_index=1,
        )


# ---- 6. row_range start ≤ header → ingestor runtime error ------------------


def test_row_range_start_inside_header_raises():
    """Phase 14.3 #6: row_range start ≤ header_row_index raises at ingest time."""
    wb = _make_workbook(
        "Sheet1",
        header_row=4,
        sections=[[("Alpha", {"Q1 2026": -100.0})]],
    )
    manifest = _manifest(
        "Sheet1",
        [{"entity_id": "ent_a", "row_range": [4, 8]}],  # start == header
        header_row_index=4,
    )
    with pytest.raises(ValueError, match="row_range start"):
        _run(wb, manifest)


# ---- 7. row_range end beyond sheet length → no crash -----------------------


def test_row_range_end_beyond_sheet_no_crash():
    """Phase 14.3 #7: row_range end past the sheet boundary emits available rows."""
    # header=row1, Alpha=row2 — only 2 rows total
    wb = _make_workbook(
        "Sheet1",
        header_row=1,
        sections=[[("Alpha", {"Q1 2026": -100.0})]],
    )
    manifest = _manifest(
        "Sheet1",
        [{"entity_id": "ent_a", "row_range": [2, 9999]}],
        header_row_index=1,
    )
    result = _run(wb, manifest)
    labels = {r.row_label for r in result.cash_flow_lines}
    assert "Alpha" in labels
    assert len(result.cash_flow_lines) == 1


# ---- 8. display_only + row_range → skipped without error -------------------


def test_display_only_with_row_range_skips_rows():
    """Phase 14.3 #8: display_only + row_range skips all rows without error."""
    wb = _make_workbook(
        "Sheet1",
        header_row=1,
        sections=[[("Alpha", {"Q1 2026": -100.0}), ("Beta", {"Q2 2026": -200.0})]],
    )
    manifest = _manifest(
        "Sheet1",
        [{"entity_id": "ent_a", "layout_type": "display_only", "row_range": [2, 3]}],
        header_row_index=1,
    )
    result = _run(wb, manifest)
    assert len(result.cash_flow_lines) == 0
    assert len(result.entities) == 1
