"""Phase 14.2 / L19 — workbook structural discovery + draft-manifest generator.

Scope: scrape the layout of an Excel workbook (read-only) and produce
a draft :class:`WorkbookManifestConfig` the human can review and
finish. The scraper discovers STRUCTURE; the manifest governs MEANING.

Discovery domain (what the scraper may infer)
=============================================

* sheet count
* per-sheet ``max_row`` / ``max_col``
* candidate header row in rows 1-15 + winning period-header format
* label column (always column A in the v7 convention)
* label-row count + subtotal-row count
* sheet role: ``family_aggregate`` / ``board_snapshot`` /
  ``assumptions_metadata`` / ``ownership_structure`` /
  ``entity_sheet`` / ``re_partnership`` / ``unknown``
* layout_type: ``horizontal_quarter`` / ``display_only`` / ``unknown``
* per-sheet confidence scores
* manifest-level majorities (header_row, period_format)

Human-classification boundary (what the scraper MUST NOT infer)
================================================================

* legal distributability
* tax treatment / withholding
* entity-governance availability
* whether OpCo cash is family-office spendable
* whether development / land value is spendable
* final ``distributable_candidate`` status
* final ``restricted`` status

These remain TODO on the draft manifest's
``row_classification_rules`` for human authoring.

Privacy posture (Phase 14 RT4 + Phase 14.2 RT)
==============================================

Two output modes:

* **privacy_safe** (default): redacts any sheet name that doesn't
  match a known structural keyword set. Suitable for a chat
  summary, a committed scaffold, or a public-eyes report.
* **local_private**: preserves real sheet names. The CLI refuses
  to write the local_private output to anything other than a
  ``*_local.yaml`` path (which is gitignored per Phase 14 +
  Phase 14.x scaffold conventions).

This module never:

* mutates the workbook
* prints live cell values
* prints row contents (label, amount, source_reference)
* commits the workbook or workbook-derived live data
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Literal

from aa_model.ingestion.schemas import (
    EntitySheetSpec,
    REPartnershipSheetSpec,
    WorkbookManifestConfig,
)
from aa_model.ingestion.workbook import (
    _is_subtotal_row,
    _parse_period_header,
)

if TYPE_CHECKING:
    pass


# ---- supported formats + scan range ----------------------------------------


_SUPPORTED_FORMATS: tuple[str, ...] = ("yyyy_q", "q_yy", "q_yyyy", "calendar_qe")
_SUBTOTAL_PATTERNS: list[str] = ["total", "subtotal", "sum", "grand total", "net cash"]
_HEADER_SCAN_ROWS: int = 15
_MIN_PARSEABLE_HEADERS_FOR_HORIZONTAL: int = 4
_MIN_LABEL_ROWS_FOR_ENTITY: int = 3


# Role-keyword classifier. Lowercase substring match; first-hit wins.
# Ordered so structural / unambiguous keywords precede entity guesses.
_ROLE_KEYWORDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    (
        "family_aggregate",
        ("summary", "family aggregate", "aggregate", "rollup", "roll-up", "cash flow", "cashflow"),
    ),
    ("board_snapshot", ("board", "snapshot", "snap")),
    ("assumptions_metadata", ("assumption", "notes", "note", "legend", "instructions")),
    ("ownership_structure", ("ownership", "structure", "org chart", "org-chart", "tree")),
    ("re_partnership", ("real estate", "real-estate", "re partnership", "partnership", " lp ")),
    ("entity_trust", ("trust", "crut", "gst", "ilit", "qprt")),
    ("entity_llc", ("llc", "holding", "holdco", "opco", "operating co")),
)


# ---- dataclasses -----------------------------------------------------------


@dataclass(frozen=True)
class HeaderCandidate:
    """One candidate header row + the format that won on it."""

    row_index: int  # 1-indexed
    period_header_format: str  # "yyyy_q" | "q_yy" | "q_yyyy" | "calendar_qe"
    parseable_count: int


@dataclass(frozen=True)
class SheetDiscovery:
    """Structural inference for a single sheet.

    ``sheet_name_raw`` carries the literal workbook sheet name. Code
    that surfaces the discovery result for chat / committed output
    must consult ``sheet_name_safe`` (which carries the placeholder
    when ``is_personal_shaped`` or non-structural).
    """

    sheet_name_raw: str
    sheet_name_safe: str
    max_row: int
    max_col: int
    role: str  # see _ROLE_KEYWORDS + "entity_sheet" / "unknown"
    role_confidence: float  # 0.0 .. 1.0
    layout_type: str  # "horizontal_quarter" | "display_only" | "unknown"
    header_candidate: HeaderCandidate | None
    label_column: int  # 1-indexed; column A by default
    label_row_count: int
    subtotal_row_count: int
    parsed_quarter_first: str | None
    parsed_quarter_last: str | None
    is_personal_shaped: bool


@dataclass
class WorkbookDiscoveryResult:
    workbook_path: str
    workbook_filename: str
    workbook_hash: str  # SHA256 hex digest of raw bytes
    total_sheets: int
    sheets: list[SheetDiscovery] = field(default_factory=list)

    # Aggregates computed at finalize time.
    role_counts: dict[str, int] = field(default_factory=dict)
    layout_counts: dict[str, int] = field(default_factory=dict)
    detected_format_majority: str | None = None
    detected_header_row_majority: int | None = None
    sheets_with_parseable_headers: int = 0
    personal_shaped_sheet_count: int = 0


@dataclass(frozen=True)
class DraftManifestResult:
    """Output of :func:`build_draft_manifest`."""

    manifest: WorkbookManifestConfig
    mode: str  # "privacy_safe" | "local_private"
    redacted_sheet_count: int
    unresolved_sheets: list[str]  # sheet_name_safe of sheets the
    # generator couldn't bucket into
    # a manifest field


# ---- privacy redaction -----------------------------------------------------


# Tokens that mark a sheet name as STRUCTURAL (privacy-safe to commit).
# Subset of _ROLE_KEYWORDS focused on names that don't carry person-
# identifying initials. We deliberately exclude "trust" / "llc" /
# "opco" / "partnership" — those are common but the abbreviations
# preceding them often decode to family-internal initials.
_STRUCTURAL_PRIVACY_KEYWORDS: tuple[str, ...] = (
    "summary",
    "family aggregate",
    "aggregate",
    "rollup",
    "roll-up",
    "cash flow",
    "cashflow",
    "board",
    "snapshot",
    "snap",
    "assumption",
    "notes",
    "legend",
    "instructions",
    "ownership",
    "structure",
    "org chart",
    "org-chart",
    "tree",
)


def _looks_personal_shaped(name: str) -> bool:
    """Conservative heuristic: classify as person-shaped if every
    non-trivial token in the sheet name is a capitalized name-shape
    (TitleCase or single-letter ALL-CAPS)."""
    tokens = [t for t in re.split(r"[^A-Za-z0-9]+", name.strip()) if t]
    if not tokens:
        return False
    # If any token matches a structural keyword family, not personal.
    lo = name.strip().lower()
    for kw in _STRUCTURAL_PRIVACY_KEYWORDS:
        if kw in lo:
            return False
    # Otherwise: all tokens must be name-shaped.
    for t in tokens:
        if not t.isalpha():
            return False  # contains digits — likely a year/code, not a name
        # TitleCase, ALLCAPS, or single-letter — name-shaped.
        if not (t.isupper() or (t[0].isupper() and (len(t) == 1 or t[1:].islower()))):
            return False
    return True


def _is_structural_safe(name: str) -> bool:
    """Return True iff the sheet name is safe to commit literally
    in privacy_safe output (matches a structural keyword)."""
    lo = name.strip().lower()
    return any(kw in lo for kw in _STRUCTURAL_PRIVACY_KEYWORDS)


def _redact_sheet_name(role: str, ordinal: int) -> str:
    """Generate a privacy-safe placeholder for a redacted sheet name."""
    bucket = role.replace("entity_", "")
    return f"<TODO_{bucket.upper()}_{ordinal}>"


# ---- header parsing helpers ------------------------------------------------


def _scan_header_candidates(
    rows: list[tuple[object, ...]],
    *,
    max_scan: int = _HEADER_SCAN_ROWS,
) -> list[HeaderCandidate]:
    """Walk rows 1..max_scan; for each, try every supported format
    and report the row + format with the highest parseable-header
    count. Returns one HeaderCandidate per row that had at least
    one parseable cell (sorted by parseable_count descending)."""
    candidates: list[HeaderCandidate] = []
    for row_idx in range(min(max_scan, len(rows))):
        row = rows[row_idx]
        if row is None:
            continue
        best_fmt: str | None = None
        best_count = 0
        for fmt in _SUPPORTED_FORMATS:
            count = sum(
                1
                for cell in row[1:]
                if cell is not None and _parse_period_header(cell, fmt) is not None
            )
            if count > best_count:
                best_count = count
                best_fmt = fmt
        if best_count > 0 and best_fmt is not None:
            candidates.append(
                HeaderCandidate(
                    row_index=row_idx + 1,
                    period_header_format=best_fmt,
                    parseable_count=best_count,
                )
            )
    candidates.sort(key=lambda c: (-c.parseable_count, c.row_index))
    return candidates


def _parsed_quarter_range(
    rows: list[tuple[object, ...]],
    candidate: HeaderCandidate,
) -> tuple[str | None, str | None]:
    """Pull the first and last parsed quarter under the winning candidate."""
    if candidate.row_index > len(rows):
        return None, None
    row = rows[candidate.row_index - 1]
    parsed: list[str] = []
    for cell in row[1:]:
        p = _parse_period_header(cell, candidate.period_header_format)
        if p is not None:
            parsed.append(p)
    if not parsed:
        return None, None
    return min(parsed), max(parsed)


# ---- role classification ---------------------------------------------------


def _classify_role_by_name(name: str) -> tuple[str, float]:
    """Match the sheet name against role keywords. Returns (role,
    confidence). Falls back to ('unknown', 0.0)."""
    lo = name.strip().lower()
    for role, keywords in _ROLE_KEYWORDS:
        for kw in keywords:
            if kw in lo:
                return role, 0.9
    return "unknown", 0.0


def _refine_role_with_shape(
    role: str,
    role_confidence: float,
    *,
    label_row_count: int,
    has_parseable_header: bool,
    max_col: int,
) -> tuple[str, float, str]:
    """Decide layout_type + adjust role using the sheet's structural shape.

    Returns (role, role_confidence, layout_type).
    """
    # Default layout: horizontal_quarter when we have a parseable
    # header AND a meaningful number of label rows. Anything else is
    # display_only (catch-all — including sparse / metadata sheets,
    # ownership graphs, summary tables that lack a quarterly grid).
    if has_parseable_header and label_row_count >= _MIN_LABEL_ROWS_FOR_ENTITY:
        layout_type = "horizontal_quarter"
    else:
        layout_type = "display_only"

    # If the role is "unknown" but the sheet shape strongly suggests
    # an entity-style cash-flow sheet, upgrade to "entity_sheet".
    if role == "unknown" and layout_type == "horizontal_quarter":
        return "entity_sheet", 0.6, layout_type

    return role, role_confidence, layout_type


# ---- discovery driver ------------------------------------------------------


def _hash_workbook_bytes(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def discover_workbook(
    workbook_path: Path | str,
) -> WorkbookDiscoveryResult:
    """Open the workbook read-only and return a structural discovery.

    The returned :class:`WorkbookDiscoveryResult` carries raw sheet
    names. Callers that surface the result publicly must transform
    via the privacy-safe naming path.

    Read-only contract identical to Phase 14:
    ``openpyxl(read_only=True, data_only=True, keep_links=False)``.
    """
    from openpyxl import load_workbook

    path = Path(workbook_path).resolve()
    if not path.is_file():
        raise FileNotFoundError(f"workbook not found at resolved path: {path}")

    workbook_hash = _hash_workbook_bytes(path)
    workbook_filename = path.name

    wb = load_workbook(
        filename=str(path),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        all_sheet_names = list(wb.sheetnames)
        result = WorkbookDiscoveryResult(
            workbook_path=str(path),
            workbook_filename=workbook_filename,
            workbook_hash=workbook_hash,
            total_sheets=len(all_sheet_names),
        )

        # Per-role ordinal counters for redaction placeholders.
        ordinal_by_role: dict[str, int] = {}

        for sheet_name in all_sheet_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            max_row = len(rows)
            max_col = max((len(r) if r else 0) for r in rows) if rows else 0

            # Header candidates.
            header_candidates = _scan_header_candidates(rows)
            best_header: HeaderCandidate | None = None
            for c in header_candidates:
                if c.parseable_count >= _MIN_PARSEABLE_HEADERS_FOR_HORIZONTAL:
                    best_header = c
                    break

            # Quarter range (under the winning header, if any).
            qf, ql = (
                _parsed_quarter_range(rows, best_header)
                if best_header is not None
                else (None, None)
            )

            # Label row + subtotal counts.
            label_row_count = 0
            subtotal_row_count = 0
            for body in rows:
                if not body:
                    continue
                cell = body[0] if body else None
                if cell is None:
                    continue
                text = str(cell).strip()
                if not text:
                    continue
                label_row_count += 1
                if _is_subtotal_row(text, _SUBTOTAL_PATTERNS):
                    subtotal_row_count += 1

            # Role + layout classification.
            role, role_conf = _classify_role_by_name(sheet_name)
            role, role_conf, layout_type = _refine_role_with_shape(
                role,
                role_conf,
                label_row_count=label_row_count,
                has_parseable_header=best_header is not None,
                max_col=max_col,
            )

            # Privacy-safe naming.
            personal = _looks_personal_shaped(sheet_name)
            structural_safe = _is_structural_safe(sheet_name)
            if personal or not structural_safe:
                bucket = role if role != "unknown" else "entity"
                ordinal_by_role[bucket] = ordinal_by_role.get(bucket, 0) + 1
                sheet_name_safe = _redact_sheet_name(bucket, ordinal_by_role[bucket])
            else:
                sheet_name_safe = sheet_name

            result.sheets.append(
                SheetDiscovery(
                    sheet_name_raw=sheet_name,
                    sheet_name_safe=sheet_name_safe,
                    max_row=max_row,
                    max_col=max_col,
                    role=role,
                    role_confidence=role_conf,
                    layout_type=layout_type,
                    header_candidate=best_header,
                    label_column=1,
                    label_row_count=label_row_count,
                    subtotal_row_count=subtotal_row_count,
                    parsed_quarter_first=qf,
                    parsed_quarter_last=ql,
                    is_personal_shaped=personal,
                )
            )
    finally:
        wb.close()

    _finalize_aggregates(result)
    return result


def _finalize_aggregates(result: WorkbookDiscoveryResult) -> None:
    """Compute the aggregate counters + manifest-level majorities."""
    role_counts: dict[str, int] = {}
    layout_counts: dict[str, int] = {}
    fmt_counts: dict[str, int] = {}
    header_row_counts: dict[int, int] = {}
    parseable_count = 0
    personal_count = 0

    for s in result.sheets:
        role_counts[s.role] = role_counts.get(s.role, 0) + 1
        layout_counts[s.layout_type] = layout_counts.get(s.layout_type, 0) + 1
        if s.is_personal_shaped:
            personal_count += 1
        if s.header_candidate is not None:
            parseable_count += 1
            fmt_counts[s.header_candidate.period_header_format] = (
                fmt_counts.get(s.header_candidate.period_header_format, 0) + 1
            )
            header_row_counts[s.header_candidate.row_index] = (
                header_row_counts.get(s.header_candidate.row_index, 0) + 1
            )

    result.role_counts = role_counts
    result.layout_counts = layout_counts
    result.sheets_with_parseable_headers = parseable_count
    result.personal_shaped_sheet_count = personal_count
    result.detected_format_majority = (
        max(fmt_counts.items(), key=lambda kv: kv[1])[0] if fmt_counts else None
    )
    result.detected_header_row_majority = (
        max(header_row_counts.items(), key=lambda kv: kv[1])[0] if header_row_counts else None
    )


# ---- draft-manifest generator ----------------------------------------------


_PrivacyMode = Literal["privacy_safe", "local_private"]


def build_draft_manifest(
    discovery: WorkbookDiscoveryResult,
    *,
    mode: _PrivacyMode = "privacy_safe",
    workbook_version: str = "v_unknown",
) -> DraftManifestResult:
    """Convert a :class:`WorkbookDiscoveryResult` into a draft
    :class:`WorkbookManifestConfig`.

    The draft is structurally complete (sheet enumeration, role
    bucketing, layout flags, manifest-level header row + format)
    but leaves all economic-classification fields as TODO:
    ``row_classification_rules`` are empty on every entity sheet.
    """
    if mode not in ("privacy_safe", "local_private"):
        raise ValueError(f"unknown mode {mode!r}")

    use_safe_name = mode == "privacy_safe"

    family_aggregate_sheets: list[str] = []
    board_snapshot_sheets: list[str] = []
    entity_sheets: list[EntitySheetSpec] = []
    re_partnership_sheets: list[REPartnershipSheetSpec] = []
    unresolved: list[str] = []
    redacted_count = 0

    # Stable entity_id assignment per role so re-running discovery on
    # the same workbook produces deterministic ids.
    ordinal_by_role: dict[str, int] = {}

    def _next_entity_id(role: str) -> str:
        ordinal_by_role[role] = ordinal_by_role.get(role, 0) + 1
        return f"{role}_{ordinal_by_role[role]:02d}"

    for s in discovery.sheets:
        sheet_name = s.sheet_name_safe if use_safe_name else s.sheet_name_raw
        if use_safe_name and s.sheet_name_safe != s.sheet_name_raw:
            redacted_count += 1

        if s.role == "family_aggregate":
            family_aggregate_sheets.append(sheet_name)
            continue
        if s.role == "board_snapshot":
            board_snapshot_sheets.append(sheet_name)
            continue

        # Decide entity_type from the inferred role.
        if s.role == "re_partnership":
            entity_type: str = "real_estate_partnership"
            cash_flow_role: str = "operating"
        elif s.role == "entity_trust":
            entity_type = "trust_family"
            cash_flow_role = "distribution_only"
        elif s.role == "entity_llc":
            entity_type = "operating_llc"
            cash_flow_role = "operating"
        elif s.role in ("assumptions_metadata", "ownership_structure"):
            entity_type = "family_aggregate"
            cash_flow_role = "operating"
        elif s.role == "entity_sheet":
            entity_type = "operating_llc"
            cash_flow_role = "operating"
        else:
            entity_type = "operating_llc"
            cash_flow_role = "operating"
            unresolved.append(sheet_name)

        # Decide layout_type — display_only when the discovery says
        # so, OR when the sheet is metadata / ownership.
        if s.layout_type == "display_only" or s.role in (
            "assumptions_metadata",
            "ownership_structure",
        ):
            layout_type = "display_only"
        elif s.layout_type == "horizontal_quarter":
            layout_type = "horizontal_quarter"
        else:
            layout_type = "display_only"  # conservative default

        entity_id = _next_entity_id(
            "entity" if s.role in ("unknown", "entity_sheet") else s.role.replace("entity_", "")
        )

        spec = EntitySheetSpec(
            sheet_name=sheet_name,
            entity_id=entity_id,
            entity_type=entity_type,  # type: ignore[arg-type]
            display_name=f"{entity_id} ({sheet_name})",
            cash_flow_role=cash_flow_role,  # type: ignore[arg-type]
            row_classification_rules=[],
            header_row_index=(s.header_candidate.row_index if s.header_candidate else None),
            period_header_format=(
                s.header_candidate.period_header_format  # type: ignore[arg-type]
                if s.header_candidate
                else None
            ),
            layout_type=layout_type,  # type: ignore[arg-type]
        )
        entity_sheets.append(spec)

    # Manifest-level defaults from the majority detections.
    default_header_row_index = (
        discovery.detected_header_row_majority
        if discovery.detected_header_row_majority is not None
        else 1
    )
    period_header_format = (
        discovery.detected_format_majority
        if discovery.detected_format_majority is not None
        else "yyyy_q"
    )

    manifest = WorkbookManifestConfig(
        workbook_version=workbook_version,
        expected_workbook_filename=discovery.workbook_filename,
        default_header_row_index=default_header_row_index,
        period_header_format=period_header_format,  # type: ignore[arg-type]
        family_aggregate_sheets=family_aggregate_sheets,
        board_snapshot_sheets=board_snapshot_sheets,
        re_partnership_sheets=re_partnership_sheets,
        entity_sheets=entity_sheets,
    )

    return DraftManifestResult(
        manifest=manifest,
        mode=mode,
        redacted_sheet_count=redacted_count,
        unresolved_sheets=sorted(unresolved),
    )


# ---- aggregate diagnostic printer (privacy-safe; no row content) ----------


def render_aggregate_diagnostics(
    discovery: WorkbookDiscoveryResult,
    draft: DraftManifestResult | None = None,
) -> str:
    """Return a formatted string of structural diagnostics suitable
    for stdout / chat. NEVER prints raw cell contents, dollar values,
    or person identifiers."""
    lines: list[str] = []
    lines.append("WORKBOOK DISCOVERY — aggregate structural diagnostics")
    lines.append("=" * 60)
    lines.append(f"workbook_filename:               {discovery.workbook_filename}")
    lines.append(f"workbook_hash (prefix):          {discovery.workbook_hash[:16]}...")
    lines.append(f"total_sheets:                    {discovery.total_sheets}")
    lines.append(f"sheets_with_parseable_headers:   {discovery.sheets_with_parseable_headers}")
    lines.append(f"detected_format_majority:        {discovery.detected_format_majority!r}")
    lines.append(f"detected_header_row_majority:    {discovery.detected_header_row_majority!r}")
    lines.append(f"personal_shaped_sheet_count:     {discovery.personal_shaped_sheet_count}")
    lines.append("")
    lines.append("role_counts:")
    for role in sorted(discovery.role_counts.keys()):
        lines.append(f"  - {role}: {discovery.role_counts[role]}")
    lines.append("")
    lines.append("layout_counts:")
    for layout in sorted(discovery.layout_counts.keys()):
        lines.append(f"  - {layout}: {discovery.layout_counts[layout]}")
    if draft is not None:
        lines.append("")
        lines.append("draft_manifest:")
        lines.append(f"  - mode:                         {draft.mode}")
        lines.append(f"  - redacted_sheet_count:         {draft.redacted_sheet_count}")
        lines.append(
            f"  - family_aggregate_sheets:      {len(draft.manifest.family_aggregate_sheets)}"
        )
        lines.append(
            f"  - board_snapshot_sheets:        {len(draft.manifest.board_snapshot_sheets)}"
        )
        lines.append(f"  - entity_sheets:                {len(draft.manifest.entity_sheets)}")
        ds = sum(1 for s in draft.manifest.entity_sheets if s.layout_type == "display_only")
        hq = sum(1 for s in draft.manifest.entity_sheets if s.layout_type == "horizontal_quarter")
        lines.append(f"    of which display_only:        {ds}")
        lines.append(f"    of which horizontal_quarter:  {hq}")
        lines.append(f"  - unresolved_sheets:            {len(draft.unresolved_sheets)}")
    return "\n".join(lines)
