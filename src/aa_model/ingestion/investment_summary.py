"""Phase 15 — Investment Summary / Account-Position ingestor.

Reads the Investment Summary workbook (read-only) against a completed
local manifest and emits normalized ``AccountRecord`` + ``PositionRecord``
lists plus ``PositionIngestionDiagnostics``.

Read-only contract
==================

* ``openpyxl(read_only=True, data_only=True, keep_links=False)``
* SHA-256 hash of raw bytes captured before opening; stored in diagnostics.
* No mutation API ever called.
* Standing CAVEAT: ``data_only=True`` reads cached formula values.

State-flow contract
===================

Pure: same workbook + same manifest → same ``PositionIngestionResult``
byte-for-byte (modulo hash, which is a function of workbook bytes).

Reviewer tightenings enforced here
====================================

* T1: ``AccountRecord`` always produced; flat ``layout_type="flat_position"``
  sheets have ``account_id`` checked; synthetic prefix allowed.
* T2: ``PositionRecord.valuation_date`` resolved via fallback chain;
  stale-valuation diagnostics computed (> 90 days before ``as_of_date``).
* T4: ``income_cash_flow_flag`` read from manifest only; never inferred.
* position_terms_status diagnostic computed per position.
"""

from __future__ import annotations

import hashlib
import math
from collections import defaultdict
from datetime import date
from pathlib import Path

from aa_model.ingestion.schemas_position import (
    _SEMI_ILLIQUID_BUCKETS,
    AccountRecord,
    AccountSheetSpec,
    ManagerTermsRecord,
    PositionIngestionDiagnostics,
    PositionIngestionResult,
    PositionManifestConfig,
    PositionRecord,
)

_STALE_THRESHOLD_DAYS = 90


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _resolve_valuation_date(
    row_date: date | None,
    spec: AccountSheetSpec,
    manifest: PositionManifestConfig,
) -> tuple[date, bool]:
    """Resolve valuation_date via T2 fallback chain.

    Returns (resolved_date, used_fallback).
    """
    if row_date is not None:
        return row_date, False
    if spec.valuation_date is not None:
        return spec.valuation_date, True
    return manifest.as_of_date, True


def _classify_terms_status(
    position: PositionRecord,
    manager_by_id: dict[str, ManagerTermsRecord],
) -> str:
    """Classify a position's manager terms completeness for diagnostics."""
    if position.manager_id is None:
        return "missing_terms"
    mgr = manager_by_id.get(position.manager_id)
    if mgr is None:
        return "missing_terms"
    if mgr.confidence == "unknown":
        return "unknown_confidence"
    # Cross-check: semi_liquid/illiquid positions should have redemption terms
    if position.liquidity_bucket in _SEMI_ILLIQUID_BUCKETS:
        has_redemption_terms = (
            mgr.notice_days is not None
            or mgr.lockup_end_date is not None
            or mgr.redemption_frequency == "none"
        )
        if not has_redemption_terms:
            return "partial_terms"
    return "complete_terms"


def ingest_investment_summary(
    workbook_path: Path,
    manifest: PositionManifestConfig,
    *,
    manifest_version: str,
) -> PositionIngestionResult:
    """Ingest the Investment Summary workbook against a completed manifest.

    Returns normalized AccountRecord + PositionRecord lists and
    PositionIngestionDiagnostics. Never mutates the workbook.
    Raises FileNotFoundError if ``workbook_path`` does not exist.
    """
    if not workbook_path.exists():
        raise FileNotFoundError(f"Investment Summary workbook not found: {workbook_path.resolve()}")

    wb_hash = _sha256(workbook_path)

    from openpyxl import load_workbook

    wb = load_workbook(
        filename=str(workbook_path),
        read_only=True,
        data_only=True,
        keep_links=False,
    )

    manager_by_id: dict[str, ManagerTermsRecord] = {m.manager_id: m for m in manifest.manager_terms}

    accounts: list[AccountRecord] = []
    positions: list[PositionRecord] = []
    unmatched_rows: list[int] = []

    try:
        for spec in manifest.accounts:
            if spec.layout_type == "display_only":
                continue

            sheet_name = spec.sheet_name
            if sheet_name not in wb.sheetnames:
                unmatched_rows.append(-1)  # sheet missing; position in sheetnames list
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Resolve effective header row (0-indexed)
            header_row_index = spec.header_row_index if spec.header_row_index is not None else 0

            # Build AccountRecord (T1: always present)
            # Use spec.valuation_date → manifest.as_of_date as account-level date
            account_date = spec.valuation_date or manifest.as_of_date
            accounts.append(
                AccountRecord(
                    account_id=spec.account_id,
                    entity_id=spec.entity_id,
                    custodian="",
                    account_type="direct",
                    valuation_date=account_date,
                    source_sheet=sheet_name,
                )
            )

            # Iterate body rows starting after the header
            for row_idx in range(header_row_index + 1, len(rows)):
                row = rows[row_idx]
                if not row:
                    unmatched_rows.append(row_idx + 1)
                    continue

                # Read market value from the designated column
                if spec.value_column_index >= len(row):
                    unmatched_rows.append(row_idx + 1)
                    continue

                raw_val = row[spec.value_column_index]
                if raw_val is None:
                    unmatched_rows.append(row_idx + 1)
                    continue

                try:
                    market_value = float(raw_val)
                except (TypeError, ValueError):
                    unmatched_rows.append(row_idx + 1)
                    continue

                if not math.isfinite(market_value) or market_value < 0:
                    unmatched_rows.append(row_idx + 1)
                    continue

                # Read optional cost basis
                cost_col = spec.position_column_mappings.get("cost_basis_usd")
                cost_basis: float | None = None
                if cost_col is not None and cost_col < len(row):
                    try:
                        cost_basis = float(row[cost_col]) if row[cost_col] is not None else None
                    except (TypeError, ValueError):
                        pass

                # Read optional unfunded commitment
                commit_col = spec.position_column_mappings.get("unfunded_commitment_usd")
                commitment: float | None = None
                if commit_col is not None and commit_col < len(row):
                    try:
                        v = row[commit_col]
                        if v is not None:
                            commitment = max(0.0, float(v))
                    except (TypeError, ValueError):
                        pass

                # Read optional manager_id
                mgr_col = spec.position_column_mappings.get("manager_id")
                manager_id: str | None = None
                if mgr_col is not None and mgr_col < len(row):
                    raw_mgr = row[mgr_col]
                    if raw_mgr is not None:
                        manager_id = str(raw_mgr).strip() or None

                # Read optional asset_class
                ac_col = spec.position_column_mappings.get("asset_class")
                asset_class = "other"
                if ac_col is not None and ac_col < len(row):
                    raw_ac = row[ac_col]
                    if raw_ac is not None:
                        asset_class = str(raw_ac).strip().lower() or "other"

                # Read optional liquidity_bucket from mappings
                lb_col = spec.position_column_mappings.get("liquidity_bucket")
                liquidity_bucket = "illiquid"
                if lb_col is not None and lb_col < len(row):
                    raw_lb = row[lb_col]
                    if raw_lb is not None:
                        liquidity_bucket = str(raw_lb).strip().lower() or "illiquid"

                # income_cash_flow_flag from manifest mapping (T4: never inferred)
                icf_col = spec.position_column_mappings.get("income_cash_flow_flag")
                income_flag = False
                if icf_col is not None and icf_col < len(row):
                    raw_icf = row[icf_col]
                    if raw_icf is not None:
                        income_flag = str(raw_icf).strip().lower() in ("true", "1", "yes")

                # T2: valuation date fallback
                vd_col = spec.position_column_mappings.get("valuation_date")
                row_date: date | None = None
                if vd_col is not None and vd_col < len(row):
                    import datetime as _dt

                    raw_vd = row[vd_col]
                    if isinstance(raw_vd, _dt.datetime):
                        row_date = raw_vd.date()
                    elif isinstance(raw_vd, date):
                        row_date = raw_vd

                valuation_date, used_fallback = _resolve_valuation_date(row_date, spec, manifest)

                position_id = f"{spec.account_id}__{row_idx + 1}"

                try:
                    pos = PositionRecord(
                        position_id=position_id,
                        account_id=spec.account_id,
                        manager_id=manager_id,
                        asset_class=asset_class,
                        market_value_usd=market_value,
                        cost_basis_usd=cost_basis,
                        unfunded_commitment_usd=commitment,
                        income_cash_flow_flag=income_flag,
                        liquidity_bucket=liquidity_bucket,
                        valuation_date=valuation_date,
                        source_row=row_idx + 1,
                    )
                except Exception:
                    unmatched_rows.append(row_idx + 1)
                    continue

                positions.append(pos)

    finally:
        wb.close()

    diag = _build_diagnostics(
        positions=positions,
        manifest=manifest,
        manifest_version=manifest_version,
        wb_hash=wb_hash,
        manager_by_id=manager_by_id,
        unmatched_rows=unmatched_rows,
    )

    return PositionIngestionResult(
        accounts=accounts,
        positions=positions,
        diagnostics=diag,
    )


def _build_diagnostics(
    positions: list[PositionRecord],
    manifest: PositionManifestConfig,
    manifest_version: str,
    wb_hash: str,
    manager_by_id: dict[str, ManagerTermsRecord],
    unmatched_rows: list[int],
) -> PositionIngestionDiagnostics:
    by_bucket: dict[str, int] = defaultdict(int)
    by_asset_class: dict[str, int] = defaultdict(int)
    unfunded_total = 0.0
    fallback_count = 0
    stale_count = 0
    max_age_days = 0

    terms_status_counts: dict[str, int] = {
        "complete_terms": 0,
        "partial_terms": 0,
        "missing_terms": 0,
        "unknown_confidence": 0,
    }
    incomplete_position_ids: list[str] = []

    for pos in positions:
        by_bucket[pos.liquidity_bucket] += 1
        by_asset_class[pos.asset_class] += 1
        if pos.unfunded_commitment_usd:
            unfunded_total += pos.unfunded_commitment_usd

        # T2: stale valuation check
        age_days = (manifest.as_of_date - pos.valuation_date).days
        if age_days > 0:
            fallback_count += 1
        if age_days > _STALE_THRESHOLD_DAYS:
            stale_count += 1
        if age_days > max_age_days:
            max_age_days = age_days

        # position_terms_status
        status = _classify_terms_status(pos, manager_by_id)
        terms_status_counts[status] = terms_status_counts.get(status, 0) + 1
        if status in ("missing_terms", "partial_terms"):
            incomplete_position_ids.append(pos.position_id)

    missing_manager = sum(1 for p in positions if p.manager_id is None)

    mgr_coverage = {m.manager_id: m.confidence for m in manifest.manager_terms}

    return PositionIngestionDiagnostics(
        workbook_hash=wb_hash,
        workbook_version=manifest.workbook_version,
        manifest_version=manifest_version,
        positions_total=len(positions),
        positions_by_bucket=dict(by_bucket),
        positions_by_asset_class=dict(by_asset_class),
        positions_missing_bucket=0,  # bucket always defaults to "illiquid"
        positions_missing_manager=missing_manager,
        unfunded_total_usd=unfunded_total,
        manager_terms_coverage=mgr_coverage,
        positions_with_incomplete_terms=incomplete_position_ids,
        position_terms_status=terms_status_counts,
        positions_with_fallback_valuation_date=fallback_count,
        stale_valuation_count=stale_count,
        max_valuation_age_days=max(max_age_days, 0),
        unmatched_rows=unmatched_rows,
    )


def load_position_manifest(path: Path | str) -> PositionManifestConfig:
    """Load and validate a PositionManifestConfig from a YAML file.

    Phase 17 / L20 reviewer tightening 1 (helper in ingestion layer, not
    orchestrator) and tightening 3 (fail fast).

    Raises:
        FileNotFoundError: if the file does not exist.
    """
    import yaml

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Position manifest not found: {p.resolve()}")
    with p.open("r", encoding="utf-8") as f:
        raw = yaml.safe_load(f)
    return PositionManifestConfig.model_validate(raw)


def render_position_report_section(result: PositionIngestionResult) -> str:
    """Render the ## Position universe (Phase 15, advisory) report section."""

    d = result.diagnostics
    manifest_version = d.manifest_version
    wb_version = d.workbook_version
    accounts_total = len(result.accounts)

    lines = [
        "## Position universe (Phase 15, advisory)",
        f"  workbook_version:    {wb_version}",
        f"  manifest_version:    {manifest_version}",
        f"  workbook_hash:       {d.workbook_hash[:16]}…",
        f"  accounts_total:      {accounts_total}",
        f"  positions_total:     {d.positions_total}",
        "",
        "  NAV by liquidity bucket:",
    ]

    bucket_order = [
        "cash_equivalent",
        "daily_liquid",
        "semi_liquid",
        "illiquid",
        "locked_strategic",
        "re_stabilized",
        "re_development",
        "re_land",
        "opco_strategic",
    ]
    for bucket in bucket_order:
        count = d.positions_by_bucket.get(bucket, 0)
        lines.append(f"    {bucket:<20} {count} position(s)")

    lines += [
        "",
        "  NAV by asset class:",
    ]
    for ac, count in sorted(d.positions_by_asset_class.items()):
        lines.append(f"    {ac:<24} {count} position(s)")

    lines += [
        "",
        f"  unfunded_commitments:        {d.unfunded_total_usd:,.0f}",
        "",
        "  Manager terms coverage:",
        f"    complete_terms:            {d.position_terms_status.get('complete_terms', 0)}",
        f"    partial_terms:             {d.position_terms_status.get('partial_terms', 0)}",
        f"    missing_terms:             {d.position_terms_status.get('missing_terms', 0)}",
        f"    unknown_confidence:        {d.position_terms_status.get('unknown_confidence', 0)}",
        "",
        f"  Stale valuations (>{_STALE_THRESHOLD_DAYS}d): {d.stale_valuation_count}",
        f"  Max valuation age:           {d.max_valuation_age_days} days",
        f"  Positions with fallback date:{d.positions_with_fallback_valuation_date}",
        "",
        f"  CAVEAT: {d.formula_cache_caveat}",
        f"  CAVEAT: {d.position_nav_caveat}",
    ]

    return "\n".join(lines)
