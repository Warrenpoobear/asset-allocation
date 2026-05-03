"""Phase 15 — Investment Summary structural discovery + draft-manifest generator.

Scope: scrape the layout of the Investment Summary workbook (read-only)
and produce a draft :class:`PositionManifestConfig` for human review.
The scraper discovers STRUCTURE; the manifest governs MEANING.

Discovery domain (what the scraper may infer)
=============================================

* sheet count and names (raw; redacted in privacy_safe mode)
* per-sheet max_row / max_col
* candidate header row (scan rows 1-8)
* candidate column indices by keyword match on header text:
  value / cost / commitment / manager / asset_class / name
* sheet role: ``account_sheet`` / ``aggregate_summary`` / ``display_only``
* detected valuation date from header text or sheet name
* income/distribution column candidates (local_private mode only — T4)

Human-classification boundary (what the scraper MUST NOT infer)
===============================================================

* legal liquidity beyond documented terms
* tax treatment / withholding
* ``income_cash_flow_flag`` (T4): proposed only in local_private mode;
  never set as the final value
* ``liquidity_bucket`` for any position
* ``manager_id`` assignment from position names
* whether manager-reported NAV is immediately accessible
* whether OpCo / RE appraised value is spendable

Privacy posture
===============

* **privacy_safe** (default): redacts sheet names containing
  non-structural keywords. Suitable for a committed scaffold or
  public-eyes output.
* **local_private**: preserves real sheet names. The CLI refuses to
  write this output to anything other than a ``*_local.yaml`` path.

This module never:
* mutates the workbook
* prints live cell values or position-level content
* commits the workbook or workbook-derived live data
"""

from __future__ import annotations

import datetime as _dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

# ---- column candidate detection --------------------------------------------

_VALUE_KEYWORDS = frozenset(
    {
        "market value",
        "mkt val",
        "nav",
        "fair value",
        "market val",
        "value",
        "balance",
        "ending balance",
        "fair mkt",
        "end market",
        "market",
        "mktvalue",
    }
)
_COST_KEYWORDS = frozenset(
    {
        "cost basis",
        "cost",
        "book value",
        "book val",
        "acquisition cost",
        "basis",
    }
)
_COMMITMENT_KEYWORDS = frozenset(
    {
        "commitment",
        "unfunded",
        "remaining commit",
        "capital commit",
        "uncalled",
    }
)
_MANAGER_KEYWORDS = frozenset(
    {
        "manager",
        "fund manager",
        "gp",
        "advisor",
        "sponsor",
        "general partner",
        "fund",
    }
)
_ASSET_CLASS_KEYWORDS = frozenset(
    {
        "asset class",
        "asset_class",
        "category",
        "type",
        "strategy",
        "sleeve",
        "class",
    }
)
_NAME_KEYWORDS = frozenset(
    {
        "name",
        "fund name",
        "account",
        "investment",
        "security",
        "description",
        "asset",
        "position",
    }
)
_INCOME_KEYWORDS = frozenset(
    {
        "distribution",
        "income",
        "dividend",
        "yield",
        "dist",
        "interest income",
    }
)

_STRUCTURAL_SHEET_KEYWORDS = frozenset(
    {
        "summary",
        "total",
        "overview",
        "aggregate",
        "allocation",
        "dashboard",
        "cover",
        "index",
        "contents",
        "instructions",
    }
)


def _match_column_role(header: str) -> str | None:
    """Return the candidate field role for a header string, or None."""
    h = header.lower().strip()
    if any(kw in h for kw in _VALUE_KEYWORDS):
        return "value_column_index"
    if any(kw in h for kw in _COST_KEYWORDS):
        return "cost_basis"
    if any(kw in h for kw in _COMMITMENT_KEYWORDS):
        return "unfunded_commitment"
    if any(kw in h for kw in _MANAGER_KEYWORDS):
        return "manager"
    if any(kw in h for kw in _ASSET_CLASS_KEYWORDS):
        return "asset_class"
    if any(kw in h for kw in _NAME_KEYWORDS):
        return "name_column_index"
    if any(kw in h for kw in _INCOME_KEYWORDS):
        return "income_candidate"
    return None


_DATE_PATTERNS = [
    re.compile(r"(?:as\s+of|as-of|asof)\s+(\w+\s+\d{1,2},?\s+\d{4})", re.IGNORECASE),
    re.compile(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"),
    re.compile(r"((?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{4})", re.IGNORECASE),
    re.compile(r"(q[1-4]\s*[-/]?\s*\d{4})", re.IGNORECASE),
]


def _detect_valuation_date_from_text(text: str) -> _dt.date | None:
    """Attempt to extract a valuation date from a header cell or sheet name."""
    for pat in _DATE_PATTERNS:
        m = pat.search(str(text))
        if m:
            raw = m.group(1).strip()
            for fmt in (
                "%B %d, %Y",
                "%B %d %Y",
                "%b %d, %Y",
                "%b %Y",
                "%B %Y",
                "%m/%d/%Y",
                "%m-%d-%Y",
                "%m/%d/%y",
            ):
                try:
                    return _dt.datetime.strptime(raw, fmt).date()
                except ValueError:
                    pass
    return None


# ---- dataclasses -----------------------------------------------------------


@dataclass
class ColumnCandidate:
    column_index: int  # 0-indexed
    header_text: str  # raw header string
    candidate_role: str  # e.g. "value_column_index", "manager", "income_candidate"


@dataclass
class SheetPositionDiscovery:
    sheet_name_raw: str  # raw name; redacted in privacy_safe output
    sheet_index: int
    role: Literal["account_sheet", "aggregate_summary", "display_only"]
    max_row: int
    max_col: int
    header_row_index: int | None
    column_candidates: list[ColumnCandidate]
    detected_valuation_date: _dt.date | None
    income_flag_candidates: list[int]  # column indices; T4 proposal only


@dataclass
class InvestmentSummaryDiscoveryResult:
    sheets: list[SheetPositionDiscovery]
    total_sheets: int = 0
    account_sheets: int = 0
    aggregate_sheets: int = 0
    display_only_sheets: int = 0
    detected_as_of_date: _dt.date | None = None


@dataclass
class DraftPositionManifestResult:
    manifest_yaml_text: str
    sheets_included: int
    income_candidates_proposed: int  # T4: proposed in local_private only


# ---- discovery -------------------------------------------------------------


def discover_investment_summary(path: Path) -> InvestmentSummaryDiscoveryResult:
    """Read-only structural scan of the Investment Summary workbook.

    Returns ``InvestmentSummaryDiscoveryResult`` without printing any
    cell values or position-level content.
    """
    from openpyxl import load_workbook

    wb = load_workbook(
        filename=str(path),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    sheet_discoveries: list[SheetPositionDiscovery] = []
    global_date: _dt.date | None = None

    try:
        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            max_row = len(rows)
            max_col = max((len(r) for r in rows if r), default=0)

            # Check sheet name for structural keywords → aggregate_summary
            name_lower = sheet_name.lower()
            is_structural = any(kw in name_lower for kw in _STRUCTURAL_SHEET_KEYWORDS)

            # Scan first 8 rows for a header candidate
            header_row_index: int | None = None
            column_candidates: list[ColumnCandidate] = []
            income_candidates: list[int] = []
            detected_date: _dt.date | None = None

            for row_idx in range(min(8, max_row)):
                row = rows[row_idx]
                if not row:
                    continue
                # Check each cell in this row for column role candidates
                roles_found: list[str] = []
                for col_idx, cell_val in enumerate(row):
                    if cell_val is None:
                        continue
                    text = str(cell_val).strip()
                    if not text:
                        continue
                    # Try date detection in first two rows
                    if row_idx < 2 and detected_date is None:
                        detected_date = _detect_valuation_date_from_text(text)
                    role = _match_column_role(text)
                    if role is not None:
                        roles_found.append(role)
                        column_candidates.append(
                            ColumnCandidate(
                                column_index=col_idx,
                                header_text=text,
                                candidate_role=role,
                            )
                        )
                        if role == "income_candidate":
                            income_candidates.append(col_idx)
                # A row with 2+ distinct column roles is a plausible header
                if header_row_index is None and len(set(roles_found)) >= 2:
                    header_row_index = row_idx

            # Classify sheet role
            has_value_col = any(c.candidate_role == "value_column_index" for c in column_candidates)
            if is_structural or max_row < 3:
                role: Literal["account_sheet", "aggregate_summary", "display_only"]
                role = "aggregate_summary" if is_structural else "display_only"
            elif has_value_col and header_row_index is not None:
                role = "account_sheet"
            else:
                role = "display_only"

            if detected_date is not None and global_date is None:
                global_date = detected_date

            sheet_discoveries.append(
                SheetPositionDiscovery(
                    sheet_name_raw=sheet_name,
                    sheet_index=sheet_index,
                    role=role,
                    max_row=max_row,
                    max_col=max_col,
                    header_row_index=header_row_index,
                    column_candidates=column_candidates,
                    detected_valuation_date=detected_date,
                    income_flag_candidates=income_candidates,
                )
            )
    finally:
        wb.close()

    account_count = sum(1 for s in sheet_discoveries if s.role == "account_sheet")
    agg_count = sum(1 for s in sheet_discoveries if s.role == "aggregate_summary")
    disp_count = sum(1 for s in sheet_discoveries if s.role == "display_only")

    return InvestmentSummaryDiscoveryResult(
        sheets=sheet_discoveries,
        total_sheets=len(sheet_discoveries),
        account_sheets=account_count,
        aggregate_sheets=agg_count,
        display_only_sheets=disp_count,
        detected_as_of_date=global_date,
    )


# ---- draft manifest builder ------------------------------------------------


def build_draft_position_manifest(
    discovery: InvestmentSummaryDiscoveryResult,
    *,
    mode: Literal["privacy_safe", "local_private"],
    workbook_version: str = "v_unknown",
    as_of_date: _dt.date | None = None,
) -> DraftPositionManifestResult:
    """Build a draft YAML manifest from discovery results.

    privacy_safe: sheet names redacted; suitable for committed scaffold.
    local_private: real sheet names preserved; CLI enforces gitignored path.
    income_cash_flow_flag proposals marked ``# PROPOSED`` (T4).
    """

    effective_date = as_of_date or discovery.detected_as_of_date
    date_str = str(effective_date) if effective_date else "<TODO_as_of_date>"

    lines: list[str] = [
        "# Investment Summary — draft manifest.",
        "# Generated by aa_model.ingestion.discover_investment_summary.",
        "# Review + complete column_mappings and manager_terms locally.",
        "# DO NOT commit this file if it carries live identifiers.",
        "#",
        "# income_cash_flow_flag proposals are marked # PROPOSED — confirm before use.",
        "#",
        "manifest_version: '1'",
        f"workbook_version: '{workbook_version}'",
        "expected_filename: '<TODO_filename>'",
        f"as_of_date: '{date_str}'",
        "accounts:",
    ]

    income_proposals_total = 0
    included = 0

    for sheet in discovery.sheets:
        if sheet.role not in ("account_sheet",):
            continue
        included += 1

        if mode == "privacy_safe":
            sheet_name_out = f"<TODO_sheet_{sheet.sheet_index}>"
            account_id_out = f"account_{sheet.sheet_index:03d}"
        else:
            sheet_name_out = sheet.sheet_name_raw
            account_id_out = f"account_{sheet.sheet_index:03d}"

        # Find best value column candidate
        value_cols = [
            c for c in sheet.column_candidates if c.candidate_role == "value_column_index"
        ]
        value_col = value_cols[0].column_index if value_cols else 1

        name_cols = [c for c in sheet.column_candidates if c.candidate_role == "name_column_index"]
        name_col = name_cols[0].column_index if name_cols else 0

        header_idx = sheet.header_row_index if sheet.header_row_index is not None else 0

        lines += [
            f"  - account_id: '{account_id_out}'",
            "    entity_id: '<TODO_entity_id>'",
            f"    sheet_name: '{sheet_name_out}'",
            "    layout_type: flat_position",
            f"    header_row_index: {header_idx}",
            f"    value_column_index: {value_col}",
            f"    name_column_index: {name_col}",
            "    position_column_mappings: {}  # <TODO: map field → column_index>",
        ]

        # Valuation date from discovery
        if sheet.detected_valuation_date:
            lines.append(f"    valuation_date: '{sheet.detected_valuation_date}'")
        else:
            lines.append("    valuation_date: null  # falls back to manifest as_of_date")

        # T4: income_cash_flow_flag proposals in local_private mode
        if mode == "local_private" and sheet.income_flag_candidates:
            income_proposals_total += len(sheet.income_flag_candidates)
            lines.append("    # PROPOSED income_cash_flow_flag columns (confirm before use):")
            for col_idx in sheet.income_flag_candidates:
                col_header = next(
                    (c.header_text for c in sheet.column_candidates if c.column_index == col_idx),
                    f"col_{col_idx}",
                )
                lines.append(f"    # income_candidate_col_{col_idx}: '{col_header}'")

    lines += [
        "manager_terms: []  # <TODO: human-authored; see ManagerTermsRecord schema>",
        "liquidity_tier_overrides: null  # optional; e.g. re_stabilized: locked_strategic",
    ]

    return DraftPositionManifestResult(
        manifest_yaml_text="\n".join(lines) + "\n",
        sheets_included=included,
        income_candidates_proposed=income_proposals_total,
    )


def render_position_diagnostics(
    discovery: InvestmentSummaryDiscoveryResult,
    draft: DraftPositionManifestResult,
) -> str:
    """Return a privacy-safe aggregate diagnostic string for stdout."""
    lines = [
        "INVESTMENT SUMMARY DISCOVERY",
        "=" * 60,
        f"  total sheets:              {discovery.total_sheets}",
        f"  account_sheets:            {discovery.account_sheets}",
        f"  aggregate_summary:         {discovery.aggregate_sheets}",
        f"  display_only:              {discovery.display_only_sheets}",
        f"  detected_as_of_date:       {discovery.detected_as_of_date or 'not detected'}",
        "",
        "DRAFT MANIFEST",
        f"  sheets_included:           {draft.sheets_included}",
        f"  income_candidates_proposed:{draft.income_candidates_proposed}",
        "",
        "No sheet names, column contents, or position values emitted.",
    ]
    return "\n".join(lines)
