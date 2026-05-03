"""Phase 14 / L19 — workbook ingestor.

Reads ``Cashflow Modeling v7.xlsx`` (or any conforming workbook) as a
read-only integration target and emits normalized entity + cash-flow
records plus diagnostics. Bridges qualifying lines into a Phase 13
``DistributionProducerConfig`` via the
:func:`workbook_lines_to_producer_config` function.

Determinism + read-only contract
================================

* Workbook opened with ``openpyxl(read_only=True, data_only=True,
  keep_links=False)``. No mutation API ever called.
* SHA256 hash of the raw .xlsx bytes captured before opening, for
  provenance + cache-key reuse.
* Phase 14 reviewer tightening 1: ``data_only=True`` reads cached
  formula values; if the workbook was edited but not recalculated /
  saved by Excel, ingested values may be stale. The standing CAVEAT
  in :class:`IngestionDiagnostics.formula_cache_caveat` is always
  surfaced.

State-flow contract
===================

Pure: same workbook + same manifest → same :class:`IngestionResult`
byte-for-byte (modulo the workbook hash, which is intentionally a
function of the input bytes). No ledger reads. No module state.
"""

from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd

from aa_model.ingestion.schemas import (
    CashFlowLineRecord,
    EntityRecord,
    EntitySheetSpec,
    IngestionDiagnostics,
    IngestionResult,
    REPartnershipSheetSpec,
    RowClassificationRule,
    WorkbookManifestConfig,
)

if TYPE_CHECKING:
    from aa_model.io.schemas import (
        DistributionEntryConfig,
        DistributionProducerConfig,
    )


# ---- period-header parsing -------------------------------------------------


_QUARTER_PATTERNS: dict[str, re.Pattern[str]] = {
    "yyyy_q": re.compile(r"^\s*(\d{4})\s*Q\s*([1-4])\s*$", re.IGNORECASE),
    "q_yy": re.compile(r"^\s*Q\s*([1-4])\s*['’]\s*(\d{2})\s*$", re.IGNORECASE),
    "q_yyyy": re.compile(r"^\s*Q\s*([1-4])\s*[-/\s]\s*(\d{4})\s*$", re.IGNORECASE),
}


def _parse_period_header(raw: object, fmt: str) -> str | None:
    """Convert a workbook column header to a canonical 'YYYY[Q1-4]'
    string. Returns None on parse failure.

    Supported formats:
      - "yyyy_q":      "2026Q1"
      - "q_yy":        "Q1'26"
      - "q_yyyy":      "Q1 2026" or "Q1-2026" or "Q1/2026"
      - "calendar_qe": a date-like cell (datetime, pandas Timestamp,
                       or quarter-end date string); converts to
                       Period(freq="Q-DEC").
    """
    if raw is None:
        return None
    if fmt == "calendar_qe":
        try:
            ts = pd.Timestamp(raw)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None
        if pd.isna(ts):
            return None
        return str(ts.to_period(freq="Q-DEC"))
    text = str(raw).strip()
    if not text:
        return None
    pat = _QUARTER_PATTERNS.get(fmt)
    if pat is None:
        return None
    m = pat.match(text)
    if not m:
        return None
    if fmt == "yyyy_q":
        year = int(m.group(1))
        quarter = int(m.group(2))
    elif fmt == "q_yy":
        quarter = int(m.group(1))
        yy = int(m.group(2))
        # Two-digit year → 20xx (workbook covers contemporary forecast
        # horizon; 19xx interpretations are out of scope).
        year = 2000 + yy
    elif fmt == "q_yyyy":
        quarter = int(m.group(1))
        year = int(m.group(2))
    else:
        return None
    return f"{year}Q{quarter}"


def _is_subtotal_row(label: str, patterns: list[str]) -> bool:
    lo = label.strip().lower()
    if not lo:
        return False
    for p in patterns:
        if p.lower() in lo:
            return True
    return False


def _classify_row(
    label: str, rules: list[RowClassificationRule]
) -> RowClassificationRule | None:
    """First-match-wins case-insensitive substring matcher."""
    if not rules:
        return None
    lo = label.strip().lower()
    for rule in rules:
        if rule.row_label_pattern.strip().lower() in lo:
            return rule
    return None


def _infer_direction_from_sign(amount_usd: float) -> str:
    return "inflow" if amount_usd >= 0.0 else "outflow"


# ---- ingestor --------------------------------------------------------------


def _hash_workbook_bytes(path: Path) -> str:
    """SHA256 hex digest of the raw .xlsx bytes — provenance key."""
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ingest_workbook(
    workbook_path: Path | str,
    manifest: WorkbookManifestConfig,
    *,
    manifest_version: str = "1",
) -> IngestionResult:
    """Read the workbook at ``workbook_path`` per ``manifest`` and
    return an :class:`IngestionResult`.

    * Read-only: opens with ``openpyxl(read_only=True, data_only=True,
      keep_links=False)``; no mutation API ever called.
    * Deterministic: same workbook bytes + same manifest →
      byte-identical IngestionResult (modulo workbook_hash, which is
      a pure function of the bytes).
    * Phase 14 reviewer tightening 3: board-snapshot reconciliation is
      ADVISORY ONLY — never raises on a delta.

    Raises:
        FileNotFoundError: if ``workbook_path`` is not an existing file.
        ValueError: on hard validation failures (required sheet missing,
            duplicate (entity_id, quarter, row_label), entity_id with
            colons, etc.).
    """
    # Local import keeps openpyxl optional at module-load time (the
    # rest of the model has no openpyxl dependency).
    from openpyxl import load_workbook

    path = Path(workbook_path).resolve()
    if not path.is_file():
        raise FileNotFoundError(
            f"workbook not found at resolved path: {path}"
        )

    workbook_hash = _hash_workbook_bytes(path)
    workbook_filename = path.name

    diag = IngestionDiagnostics(
        workbook_hash=workbook_hash,
        workbook_filename=workbook_filename,
        workbook_version=manifest.workbook_version,
        manifest_version=manifest_version,
    )
    result = IngestionResult(diagnostics=diag)

    wb = load_workbook(
        filename=str(path),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        all_sheet_names = list(wb.sheetnames)

        # Build the role index up front so we can detect missing
        # required sheets and unmapped sheets cleanly.
        entity_specs: dict[str, EntitySheetSpec] = {
            s.sheet_name: s for s in manifest.entity_sheets
        }
        re_specs: dict[str, REPartnershipSheetSpec] = {
            s.sheet_name: s for s in manifest.re_partnership_sheets
        }
        family_aggregate = set(manifest.family_aggregate_sheets)
        board_snapshots = set(manifest.board_snapshot_sheets)

        all_mapped = (
            set(entity_specs)
            | set(re_specs)
            | family_aggregate
            | board_snapshots
        )

        # Required sheets: every entity_sheet, every re_partnership_sheet,
        # every family_aggregate_sheet, every board_snapshot_sheet must
        # be present. Missing → hard error.
        required = (
            set(entity_specs)
            | set(re_specs)
            | family_aggregate
            | board_snapshots
        )
        missing_required = sorted(required - set(all_sheet_names))
        if missing_required:
            raise ValueError(
                f"workbook missing required sheet(s) declared in manifest: "
                f"{missing_required}"
            )

        # Unmapped: present in the workbook but not declared anywhere
        # in the manifest. Surfaced as INFO; ingestion proceeds.
        diag.unmapped_sheets = sorted(set(all_sheet_names) - all_mapped)

        # Entity-sheet ingestion.
        for sheet_name, spec in entity_specs.items():
            _ingest_entity_sheet(
                wb=wb,
                sheet_name=sheet_name,
                spec=spec,
                manifest=manifest,
                workbook_filename=workbook_filename,
                result=result,
            )
            diag.sheets_ingested.append(sheet_name)

        # RE-partnership ingestion (same parser plus per-row asset_id
        # assignment).
        for sheet_name, spec in re_specs.items():
            _ingest_entity_sheet(
                wb=wb,
                sheet_name=sheet_name,
                spec=spec,
                manifest=manifest,
                workbook_filename=workbook_filename,
                result=result,
            )
            diag.sheets_ingested.append(sheet_name)

        # Family-aggregate sheets are parsed for reconciliation only —
        # totals are extracted into board_snapshot_reconciliations.
        # Phase 14 reviewer tightening 3: ADVISORY ONLY.
        for sheet_name in sorted(family_aggregate):
            _reconcile_aggregate_sheet(
                wb=wb,
                sheet_name=sheet_name,
                manifest=manifest,
                result=result,
            )
            diag.sheets_ingested.append(sheet_name)

        # Board-snapshot sheets: same advisory-reconciliation treatment.
        for sheet_name in sorted(board_snapshots):
            _reconcile_aggregate_sheet(
                wb=wb,
                sheet_name=sheet_name,
                manifest=manifest,
                result=result,
            )
            diag.sheets_ingested.append(sheet_name)

        _finalize_diagnostics(result)
    finally:
        wb.close()

    return result


def _ingest_entity_sheet(
    *,
    wb: object,
    sheet_name: str,
    spec: EntitySheetSpec,
    manifest: WorkbookManifestConfig,
    workbook_filename: str,
    result: IngestionResult,
) -> None:
    """Parse one entity sheet into entities + cash-flow line rows.

    Phase 14.1 layout knobs:
      * spec.header_row_index → manifest.default_header_row_index → 1
      * spec.period_header_format → manifest.period_header_format
      * spec.layout_type == "display_only" → declare entity, skip rows
    """
    ws = wb[sheet_name]  # type: ignore[index]

    # Construct the EntityRecord first so the ingestor surfaces a
    # clear schema-validation error on bad manifest input.
    entity = EntityRecord(
        entity_id=spec.entity_id,
        display_name=spec.display_name,
        entity_type=spec.entity_type,
        parent_entity_id=spec.parent_entity_id,
        cash_flow_role=spec.cash_flow_role,
        source_sheet=sheet_name,
        source_workbook=workbook_filename,
    )
    result.entities.append(entity)

    # Phase 14.1: display_only sheets are declared but their data rows
    # are NOT extracted. Used for sheets that are entity-shaped but
    # are display/summary tables, not data tables (e.g., board
    # snapshots when declared as entity sheets, ownership graphs).
    # The unmapped-sheets diagnostic gets a sibling tracker for these.
    if spec.layout_type == "display_only":
        result.diagnostics.missing_optional_sheets.append(
            f"display_only:{sheet_name}"
        )
        return

    # Resolve effective layout knobs from spec → manifest defaults.
    header_row_index = (
        spec.header_row_index
        if spec.header_row_index is not None
        else manifest.default_header_row_index
    )
    period_header_format = (
        spec.period_header_format
        if spec.period_header_format is not None
        else manifest.period_header_format
    )

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    if header_row_index > len(rows):
        # Sheet too short for the configured header row; nothing to parse.
        return

    header = rows[header_row_index - 1]
    if not header or len(header) < 2:
        return

    # Map column index → canonical quarter string (ignoring columns
    # whose header doesn't parse).
    column_quarters: dict[int, str] = {}
    for col_idx, raw_header in enumerate(header[1:], start=1):
        canonical = _parse_period_header(raw_header, period_header_format)
        if canonical is not None:
            column_quarters[col_idx] = canonical
        else:
            if raw_header is not None and str(raw_header).strip():
                # Track unparseable headers for diagnostics; skip the column.
                result.diagnostics.unparseable_period_headers.append(
                    f"{sheet_name}!col{col_idx}: {raw_header!r}"
                )

    # asset_id_by_row_label is REPartnershipSheetSpec-only.
    asset_id_by_row_label: dict[str, str] = {}
    if isinstance(spec, REPartnershipSheetSpec):
        asset_id_by_row_label = {
            label.strip().lower(): aid
            for label, aid in spec.asset_id_by_row_label.items()
        }

    # Track (entity_id, quarter, row_label) to detect duplicates.
    seen_keys: set[tuple[str, str, str]] = set()

    # Body rows start AFTER the header row (Phase 14.1: header_row_index
    # is configurable, so the body slice + 1-indexed row_idx must shift
    # accordingly).
    body_start = header_row_index  # 1-indexed = row immediately after header
    for row_idx, body in enumerate(rows[body_start:], start=body_start + 1):
        if not body or len(body) < 2:
            result.diagnostics.blank_rows_skipped += 1
            continue
        raw_label = body[0]
        if raw_label is None:
            # Wholly blank row.
            result.diagnostics.blank_rows_skipped += 1
            continue
        label = str(raw_label).strip()
        if not label:
            result.diagnostics.blank_rows_skipped += 1
            continue

        if _is_subtotal_row(label, manifest.subtotal_label_patterns):
            result.diagnostics.excluded_subtotal_rows += 1
            continue

        rule = _classify_row(label, spec.row_classification_rules)

        for col_idx, quarter in column_quarters.items():
            if col_idx >= len(body):
                continue
            raw_amount = body[col_idx]
            if raw_amount is None:
                # Empty cell — not a data point; not counted as blank-row.
                continue
            try:
                amount = float(raw_amount)
            except (TypeError, ValueError):
                continue
            if amount == 0.0:
                # Zero is structurally not a flow; suppress.
                continue

            key = (spec.entity_id, quarter, label)
            if key in seen_keys:
                # Phase 14.1 privacy fix: do NOT include the raw row
                # label in the error message — the label may contain
                # entity names, dollar amounts, or transaction-level
                # detail. Surface position + label length only.
                raise ValueError(
                    f"duplicate (entity_id, quarter, row_label) at "
                    f"{sheet_name}!row{row_idx} col{col_idx}: "
                    f"entity_id={spec.entity_id!r}, "
                    f"quarter={quarter!r}, "
                    f"row_label_length={len(label)} (content redacted). "
                    f"If the sheet legitimately contains repeated row "
                    f"labels (typical of aggregate / summary sheets), "
                    f"declare it with layout_type='display_only'."
                )
            seen_keys.add(key)

            if rule is not None:
                # The rule supplies the strict classification but the
                # ingestor still enforces sign-convention consistency.
                # If the rule declares direction='inflow' but the
                # workbook cell is negative, surface as an unmatched
                # line rather than mint a CashFlowLineRecord that
                # would fail validation.
                expected_sign_inflow = rule.direction == "inflow"
                # Phase 14.1 privacy fix: do NOT include the raw row
                # label or amount in the unmatched-sample entry. The
                # sample is rendered into the report; raw labels +
                # amounts can leak entity / dollar / transaction
                # detail. Surface position + sign-mismatch class only.
                if expected_sign_inflow and amount < 0:
                    result.diagnostics.unmatched_lines_count += 1
                    if len(result.diagnostics.unmatched_lines_sample) < 8:
                        result.diagnostics.unmatched_lines_sample.append(
                            f"{sheet_name}!row{row_idx} ({quarter}): "
                            f"sign mismatch (rule expected inflow; got negative)"
                        )
                    continue
                if not expected_sign_inflow and amount > 0:
                    result.diagnostics.unmatched_lines_count += 1
                    if len(result.diagnostics.unmatched_lines_sample) < 8:
                        result.diagnostics.unmatched_lines_sample.append(
                            f"{sheet_name}!row{row_idx} ({quarter}): "
                            f"sign mismatch (rule expected outflow; got positive)"
                        )
                    continue
                line = CashFlowLineRecord(
                    source_workbook=workbook_filename,
                    sheet_name=sheet_name,
                    row_label=label,
                    entity_id=spec.entity_id,
                    quarter=quarter,
                    amount_usd=amount,
                    category=rule.category,
                    direction=rule.direction,
                    certainty=rule.certainty,
                    recurrence_type=rule.recurrence_type,
                    distributable_candidate=rule.distributable_candidate,
                    restricted=rule.restricted,
                    source_reference=f"{sheet_name}!{label}",
                )
            else:
                # Unclassified: emit with default classification so the
                # broader entity / cash-flow table is complete, but the
                # producer bridge won't pick it up (distributable_candidate=False).
                direction = _infer_direction_from_sign(amount)
                line = CashFlowLineRecord(
                    source_workbook=workbook_filename,
                    sheet_name=sheet_name,
                    row_label=label,
                    entity_id=spec.entity_id,
                    quarter=quarter,
                    amount_usd=amount,
                    category="unknown",
                    direction=direction,
                    certainty="forecast",
                    recurrence_type="unknown",
                    distributable_candidate=False,
                    restricted=False,
                    source_reference=f"{sheet_name}!{label}",
                )
                result.diagnostics.unmatched_lines_count += 1
                if len(result.diagnostics.unmatched_lines_sample) < 8:
                    # Phase 14.1 privacy fix: row-position only; no
                    # label content. Renders into the report.
                    result.diagnostics.unmatched_lines_sample.append(
                        f"{sheet_name}!row{row_idx} ({quarter}): unclassified"
                    )

            # Tag with asset_id (RE partnership case) for downstream
            # bridge use. We piggyback on source_reference so the
            # ledger-row schema doesn't need a new column.
            if asset_id_by_row_label:
                aid = asset_id_by_row_label.get(label.strip().lower())
                if aid is not None:
                    line = line.model_copy(update={
                        "source_reference": f"asset_id={aid}|{line.source_reference}",
                    })

            result.cash_flow_lines.append(line)


def _reconcile_aggregate_sheet(
    *,
    wb: object,
    sheet_name: str,
    manifest: WorkbookManifestConfig,
    result: IngestionResult,
) -> None:
    """Phase 14 reviewer tightening 3: ADVISORY ONLY reconciliation.

    Family-aggregate / board-snapshot sheets carry roll-up totals the
    workbook author maintains. The ingestor reads them, recomputes
    the ingestor-side total over the same horizon, and surfaces any
    delta as a diagnostic entry. Never raises on a delta.

    Layout convention: same as entity sheets (column A row labels,
    columns B+ quarter headers). The ingestor sums all numeric rows
    that aren't subtotals; the resulting total is compared to the
    explicit subtotal/total row if present (or zero).
    """
    ws = wb[sheet_name]  # type: ignore[index]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    # Phase 14.1: aggregate sheets use the manifest-level
    # default_header_row_index. Per-sheet override would require
    # promoting board_snapshot_sheets / family_aggregate_sheets to
    # structured specs — out of scope for Phase 14.1.
    header_row_index = manifest.default_header_row_index
    if header_row_index > len(rows):
        return
    header = rows[header_row_index - 1]
    if not header or len(header) < 2:
        return

    column_quarters: dict[int, str] = {}
    for col_idx, raw_header in enumerate(header[1:], start=1):
        canonical = _parse_period_header(raw_header, manifest.period_header_format)
        if canonical is not None:
            column_quarters[col_idx] = canonical

    snapshot_total = 0.0
    detail_total = 0.0
    body_start = header_row_index
    for body in rows[body_start:]:
        if not body or len(body) < 2:
            continue
        raw_label = body[0]
        if raw_label is None:
            continue
        label = str(raw_label).strip()
        if not label:
            continue
        is_subtotal = _is_subtotal_row(label, manifest.subtotal_label_patterns)
        for col_idx in column_quarters:
            if col_idx >= len(body):
                continue
            raw_amount = body[col_idx]
            if raw_amount is None:
                continue
            try:
                amount = float(raw_amount)
            except (TypeError, ValueError):
                continue
            if is_subtotal:
                snapshot_total += amount
            else:
                detail_total += amount

    abs_delta = abs(snapshot_total - detail_total)
    abs_delta_pct = (
        abs_delta / abs(snapshot_total) * 100.0 if snapshot_total != 0.0 else 0.0
    )
    result.diagnostics.board_snapshot_reconciliations.append(
        (
            sheet_name,
            float(snapshot_total),
            float(detail_total),
            float(abs_delta),
            float(abs_delta_pct),
        )
    )


def _finalize_diagnostics(result: IngestionResult) -> None:
    """Compute per-entity totals + distribution-candidate breakdown."""
    diag = result.diagnostics

    inflows: dict[str, float] = {}
    outflows: dict[str, float] = {}
    for line in result.cash_flow_lines:
        if line.direction == "inflow":
            inflows[line.entity_id] = inflows.get(line.entity_id, 0.0) + line.amount_usd
        else:
            outflows[line.entity_id] = (
                outflows.get(line.entity_id, 0.0) + line.amount_usd
            )
    diag.total_inflows_usd_by_entity = {k: float(v) for k, v in sorted(inflows.items())}
    diag.total_outflows_usd_by_entity = {
        k: float(v) for k, v in sorted(outflows.items())
    }

    candidates_by_domain: dict[str, float] = {}
    cand_count = 0
    excl_restricted_count = 0
    excl_restricted_usd = 0.0
    for line in result.cash_flow_lines:
        if not line.distributable_candidate:
            continue
        if line.restricted:
            excl_restricted_count += 1
            excl_restricted_usd += line.amount_usd
            continue
        # Domain comes from the matched RowClassificationRule; the
        # bridge re-derives it from the line's category convention or
        # carried-along source_reference. For diagnostics we rely on
        # the bridge to populate domain rollups; here we only count.
        cand_count += 1

    diag.distribution_candidates_count = cand_count
    diag.excluded_restricted_count = excl_restricted_count
    diag.excluded_restricted_usd = float(excl_restricted_usd)
    diag.distribution_candidates_by_domain_usd = candidates_by_domain


# ---- workbook → producer bridge --------------------------------------------


def workbook_lines_to_producer_config(
    result: IngestionResult,
    manifest: WorkbookManifestConfig,
) -> DistributionProducerConfig:
    """Phase 14 → Phase 13 bridge.

    Convert qualifying workbook lines into a
    :class:`aa_model.io.schemas.DistributionProducerConfig` consumable
    by the existing Phase 13 producer ABC.

    Inclusion rules:
      * line.distributable_candidate == True
      * line.restricted == False
      * line.direction == "inflow"
      * line.amount_usd > 0
      * line.recurrence_type ∈ {"recurring", "one_time"}  (not "unknown")
      * the matched row-classification rule supplied a ``domain``
        (enforced at manifest validation time)

    Phase 14 reviewer tightening 2: producer_id uses
    ``workbook_version`` (NOT ``workbook_hash``) so cross-run audit
    is stable across workbook edits within the same version.
      producer_id = f"{workbook_version}__{sheet}__{row_label}__{quarter}"
    """
    # Local import to avoid a top-level cycle: io/schemas imports
    # nothing from ingestion/, but ingestion/ depends on the Phase 13
    # producer-config schema at bridge time.
    from aa_model.io.schemas import (
        DistributionEntryConfig,
        DistributionProducerConfig,
    )

    # Reverse-lookup table: sheet_name → entity-spec object so we can
    # pull domain hints for non-RE-partnership entities and asset_id
    # mappings for RE-partnership rows.
    spec_by_sheet: dict[str, EntitySheetSpec] = {}
    for s in manifest.entity_sheets:
        spec_by_sheet[s.sheet_name] = s
    for s in manifest.re_partnership_sheets:
        spec_by_sheet[s.sheet_name] = s

    domain_by_domain_count: dict[str, float] = {}
    entries: list[DistributionEntryConfig] = []
    for line in result.cash_flow_lines:
        if not line.distributable_candidate:
            continue
        if line.restricted:
            continue
        if line.direction != "inflow":
            continue
        if line.amount_usd <= 0.0:
            continue
        if line.recurrence_type not in ("recurring", "one_time"):
            continue

        spec = spec_by_sheet.get(line.sheet_name)
        if spec is None:
            continue
        # Re-derive the matched rule to get domain. This is a re-match
        # since the line itself doesn't carry the rule pointer — but
        # the manifest is small (handful of rules per sheet) and
        # ingestion is called once per run.
        rule = _classify_row(line.row_label, spec.row_classification_rules)
        if rule is None or rule.domain is None:
            continue

        # asset_id for RE-partnership rows is encoded in source_reference
        # as "asset_id=<id>|...".
        asset_id: str | None = None
        if line.source_reference and line.source_reference.startswith("asset_id="):
            head = line.source_reference.split("|", 1)[0]
            asset_id = head.removeprefix("asset_id=")

        producer_id = (
            f"{manifest.workbook_version}__"
            f"{line.sheet_name}__"
            f"{line.row_label}__"
            f"{line.quarter}"
        )
        # Sanitize colons in producer_id: sheet_name and row_label may
        # legitimately contain spaces/punctuation, but a colon would
        # break the Phase 12.5 source convention. Replace defensively.
        producer_id = producer_id.replace(":", "_")
        entries.append(
            DistributionEntryConfig(
                producer_id=producer_id,
                domain=rule.domain,
                entity_id=line.entity_id,
                asset_id=asset_id,
                quarter=line.quarter,
                amount_usd=line.amount_usd,
                recurrence_type=line.recurrence_type,
                confidence=line.certainty,
                restricted=False,
                source_reference=f"workbook={result.diagnostics.workbook_filename}|"
                                 f"sheet={line.sheet_name}|row={line.row_label}",
            )
        )
        domain_by_domain_count[rule.domain] = (
            domain_by_domain_count.get(rule.domain, 0.0) + line.amount_usd
        )

    # Update the diagnostics' by-domain rollup now that we know it.
    result.diagnostics.distribution_candidates_by_domain_usd = domain_by_domain_count

    return DistributionProducerConfig(entries=entries)
